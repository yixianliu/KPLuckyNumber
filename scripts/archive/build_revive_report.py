# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
生成「自学习闭环复活」报告（Excel）
====================================
读取:
  - reports/diagnostic/revive_loop.json      (EWMA 权重对比 + 写库统计)
  - reports/diagnostic/v312_production.json  (本轮 EWMA 自适应权重命中率)
对比基线（首轮, weight_history=0 时测得的硬编码默认权重命中率，来源标注）:
  - Top-1=11.5% / Top-3=32.67% / Top-5=50.83% / Top-6=60.33% / match=3.017

配色: 蓝=输入 黑=计算 绿=增益/超基线 红=减益/低于基线
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

SRC_LOOP = 'reports/diagnostic/revive_loop.json'
SRC_VERIFY = 'reports/diagnostic/v312_production.json'
OUT = 'reports/diagnostic/排列5_自学习闭环复活报告.xlsx'

BLUE = '1F4E79'; BLACK = '000000'; GREEN = '2E7D32'; RED = 'C62828'; WHITE = 'FFFFFF'
HEADER = '1F4E79'
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(wrap_text=True, vertical='top')


def font(color=BLACK, bold=False, size=10):
    return Font(name='微软雅黑', color=color, bold=bold, size=size)


def fill(c):
    return PatternFill('solid', fgColor=c)


def put(ws, r, c, v, color=BLACK, bold=False, fmt=None, align=None, fillc=None, size=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font(color, bold, size)
    if fmt:
        cell.number_format = fmt
    cell.alignment = align or CENTER
    cell.border = BORDER
    if fillc:
        cell.fill = fill(fillc)
    return cell


def hdr(ws, r, cols):
    for i, c in enumerate(cols, 1):
        put(ws, r, i, c, WHITE, True, fillc=HEADER)


def build():
    loop = json.load(open(SRC_LOOP, encoding='utf-8'))
    verify = json.load(open(SRC_VERIFY, encoding='utf-8'))

    default_w = loop['default_weights']
    ewma = loop['ewma_after_replay']
    algos = list(default_w.keys())
    ALGO_CN = {'frequency_weighted': '频率加权', 'omission_regression': '遗漏回归',
               'trend_momentum': '趋势动量', 'markov_transition': '马尔可夫',
               'pattern_continuation': '形态延续', 'bayesian_inference': '贝叶斯推断',
               'feature_engineering': '特征工程'}

    HARD = {'top1': 11.5, 'top3': 32.67, 'top5': 50.83, 'top6': 60.33, 'match': 3.017}
    RB = {'top1': 10.0, 'top3': 30.0, 'top5': 50.0, 'top6': 60.0}
    EW = {'top1': verify['top1_overall'], 'top3': verify['top3_overall'],
          'top5': verify['top5_overall'], 'top6': verify['top6_overall_rate'],
          'match': verify['avg_top6_match_count']}
    # 预计算标量，避免 f-string 内嵌字典引号
    h_t1, e_t1, b_t1 = HARD['top1'], EW['top1'], RB['top1']
    h_t3, e_t3, b_t3 = HARD['top3'], EW['top3'], RB['top3']
    h_t5, e_t5, b_t5 = HARD['top5'], EW['top5'], RB['top5']
    h_t6, e_t6, b_t6 = HARD['top6'], EW['top6'], RB['top6']
    h_m, e_m = HARD['match'], EW['match']
    d_t6, d_t5 = e_t6 - h_t6, e_t5 - h_t5
    d_t1 = h_t1 - e_t1

    wb = Workbook()

    # ---- Sheet1 概览 ----
    ws = wb.active
    ws.title = '概览'
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 50
    put(ws, 1, 1, '自学习闭环复活报告', BLUE, True, size=14, align=Alignment(vertical='center'))
    ws.merge_cells('A1:C1')
    rows = [
        ('指标', '数值', '说明'),
        ('weight_history 产物条数', '0 -> ' + str(loop['written_art']),
         '复活前为0（死路径），现已积累真实验证'),
        ('p5_prediction_record 新增', loop['written_pred'],
         "新 report_uuid=v3.12-verify-loop，不污染原有992条"),
        ('enable_adaptive_weights', 'True (默认)',
         '下次任何预测自动回放 EWMA 并融合进权重'),
        ('验证窗口', '最近 ' + str(loop['tested']) + ' 期真实已开奖',
         'walk-forward 防前视偏差，纯算法关AI'),
        ('闭环状态', '已复活且生效', '数据积累 + 权重自适应双通路打通'),
    ]
    r = 3
    hdr(ws, r, rows[0])
    for name, val, desc in rows[1:]:
        r += 1
        put(ws, r, 1, name, BLUE)
        put(ws, r, 2, val, BLACK, bold=True)
        put(ws, r, 3, desc, BLACK, align=WRAP)

    # ---- Sheet2 EWMA 权重对比 ----
    ws2 = wb.create_sheet('EWMA权重对比')
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    hdr(ws2, 1, ['算法', '默认v3.12权重', '回放后EWMA', '变化'])
    r = 2
    for a in algos:
        dv = default_w[a]
        ev = ewma[a]
        delta = ev - dv
        put(ws2, r, 1, ALGO_CN[a], BLUE)
        put(ws2, r, 2, round(dv, 4), BLACK, fmt='0.0000')
        put(ws2, r, 3, round(ev, 4), BLACK, fmt='0.0000')
        put(ws2, r, 4, round(delta, 4), GREEN if delta > 0 else RED, fmt='+0.0000;-0.0000')
        r += 1
    chart = BarChart()
    chart.title = '默认权重 vs 回放后EWMA'
    chart.type = 'col'
    chart.height = 9
    chart.width = 18
    data = Reference(ws2, min_col=2, max_col=3, min_row=1, max_row=r - 1)
    cats = Reference(ws2, min_col=1, max_col=1, min_row=2, max_row=r - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = BLUE
    chart.series[1].graphicalProperties.solidFill = '888888'
    ws2.add_chart(chart, 'A' + str(r + 3))

    # ---- Sheet3 命中率对比 ----
    ws3 = wb.create_sheet('命中率对比')
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 14
    hdr(ws3, 1, ['口径', '硬编码默认权重', 'EWMA自适应权重', '随机基线'])
    comp = [
        ('整体 Top-1', h_t1, e_t1, b_t1),
        ('整体 Top-3', h_t3, e_t3, b_t3),
        ('整体 Top-5', h_t5, e_t5, b_t5),
        ('整体 Top-6(生产)', h_t6, e_t6, b_t6),
        ('平均match_count/5', h_m, e_m, 3.00),
    ]
    r = 2
    for name, h, e, b in comp:
        put(ws3, r, 1, name, BLUE)
        put(ws3, r, 2, h, BLACK, fmt='0.00"%"' if '率' in name else '0.000')
        put(ws3, r, 3, e, BLACK, fmt='0.00"%"' if '率' in name else '0.000')
        put(ws3, r, 4, b, BLACK, fmt='0.00"%"' if '率' in name else '0.000')
        gain = e > b
        put(ws3, r, 4, ('Y ' if gain else 'N ') + 'base ' + str(b), GREEN if gain else RED, bold=True)
        r += 1
    put(ws3, r + 1, 1, '注: 硬编码基线=首轮验证(weight_history=0时, 自适应未生效)测得; '
        'EWMA=本轮回放120期后重测。来源: revive_loop.json + v312_production.json。',
        BLACK, size=8, align=WRAP)
    ws3.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)

    # ---- Sheet4 结论与建议 ----
    ws4 = wb.create_sheet('结论与建议')
    ws4.column_dimensions['A'].width = 105
    L = []
    L.append(('自学习闭环复活 - 让数字说话', BLUE, True, 13))
    L.append((('1. 闭环已复活: weight_history 从 0 到 %d 条; '
               'p5_prediction_record 新增 %d 条(新 uuid 不污染旧数据)。')
              % (loop['written_art'], loop['written_pred']), BLACK, False, 10))
    L.append(('2. enable_adaptive_weights 默认 True: 下次任何预测自动回放 EWMA 并融合权重, '
              '闭环「数据积累 + 权重自适应」双通路打通。', BLACK, False, 10))
    L.append((('3. EWMA 自适应效果(覆盖类指标增益): Top-6 %.2f%%->%.2f%%(+%.2fpp), '
               'Top-5 %.2f%%->%.2f%%(+%.2fpp)。')
              % (h_t6, e_t6, d_t6, h_t5, e_t5, d_t5), GREEN, False, 10))
    L.append((('4. 代价(精准类略降): Top-1 %.2f%%->%.2f%%(-%.2fpp)。'
               '原因: 所有算法真实命中率约随机, EWMA 把权重拉向更均匀, 牺牲频率集中的 Top-1 精准度。')
              % (h_t1, e_t1, d_t1), RED, False, 10))
    L.append(('5. 系统学到的事实: 7 算法在真实数据上都约随机(per-algo Top-5 命中率 0.4~0.65), '
              '自适应权重收敛于近似等权 - 这是诚实的数据驱动结论, 非缺陷。', BLACK, False, 10))
    L.append(('', BLACK, False, 6))
    L.append(('所以呢？下一步动作', BLUE, True, 12))
    L.append(('A. 持续积累: 让验证闭环逐期运行(每期开奖->预测->验证->写 weight_history), '
              '目标 >300 期让 EWMA 更稳健, 减少短期波动。', BLACK, False, 10))
    L.append(('B. 调参防过均匀: 若想保留频率/遗漏优势, 可限制次要算法权重上限, '
              '或调小 EWMA alpha(让历史默认权重保持更大话语权)。', BLACK, False, 10))
    L.append(('C. 产品定位: 系统目前是覆盖率/缩号工具(Top-6约62%), 非精准命中预测, 对外表述如实。', BLACK, False, 10))
    L.append(('D. 换信号源: 要真正突破随机天花板, 需引入新特征/外部信号, 继续调权重收益有限。', BLACK, False, 10))
    L.append(('', BLACK, False, 6))
    L.append(('数据来源: revive_loop.json + v312_production.json; 生成时间 2026-07-18; '
              '验证窗口最近%d期真实开奖。' % loop['tested'], BLACK, False, 8))
    r = 1
    for text, color, bold, size in L:
        put(ws4, r, 1, text, color, bold, size=size, align=WRAP)
        r += 1

    os.makedirs('reports/diagnostic', exist_ok=True)
    wb.save(OUT)
    print('已生成 ' + OUT + ' (sheets: ' + str(wb.sheetnames) + ')')


if __name__ == '__main__':
    build()
