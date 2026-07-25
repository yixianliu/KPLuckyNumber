# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""生成《排列5预测记录数据质量红线修复报告.xlsx》"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- 配色规范 (析数数) ----
C_TITLE = "1F4E79"   # 深蓝标题
C_HEAD = "2E75B6"    # 表头蓝
C_INPUT = "BDD7EE"   # 蓝底=硬编码输入
C_ASSUME = "FFF2CC"  # 黄底=重要假设
C_OK = "C6EFCE"      # 绿=通过
C_BAD = "FFC7CE"     # 红=问题
C_BLACK = "000000"
WHITE = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFONT = Font(bold=True, color=WHITE, size=11)
TFONT = Font(bold=True, color=WHITE, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(horizontal="center", vertical="center")

wb = Workbook()

def style_header(ws, row, ncol, fill=C_HEAD):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = HFONT
        cell.alignment = CEN
        cell.border = BORDER

def put(ws, r, c, v, fill=None, bold=False, color=C_BLACK, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if bold:
        cell.font = Font(bold=True, color=color)
    elif color != C_BLACK:
        cell.font = Font(color=color)
    if align:
        cell.alignment = align
    if border:
        cell.border = BORDER
    return cell

# =====================================================================
# Sheet 1: 概览
# =====================================================================
ws = wb.active
ws.title = "概览"
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 40

ws.merge_cells("A1:D1")
put(ws, 1, 1, "排列5 预测记录 · 数据质量红线修复报告", fill=C_TITLE, bold=True, color=WHITE)
ws.cell(1, 1).font = TFONT
ws.cell(1, 1).alignment = CEN
ws.row_dimensions[1].height = 28

put(ws, 2, 1, "修复日期", bold=True); put(ws, 2, 2, "2026-07-19", fill=C_INPUT)
put(ws, 2, 3, "数据表", bold=True); put(ws, 2, 4, "lucky_number.p5_prediction_record")
put(ws, 3, 1, "总记录数", bold=True); 
tot = put(ws, 3, 2, 1117, fill=C_INPUT)
put(ws, 3, 4, "来源: opt_dq_inspect.py 全表统计 (可追溯)", align=WRAP)

put(ws, 5, 1, "核心结论", fill=C_HEAD, bold=True, color=WHITE)
ws.cell(5,1).font = HFONT
ws.merge_cells("A5:D5")
rows = [
    ("修复前红线违规", "11 条嵌套旧格式 + 1 条全空 = 12 处", "现代读取器外(Excel/通用脚本)不可解析"),
    ("修复动作", "全表备份 + 嵌套→扁平 + 全空隔离", "最小变更, 不动已存命中统计"),
    ("修复后红线违规", "0 处 (PASS ✅)", "1117 条全部扁平可读"),
    ("回滚保障", "p5_prediction_record_bak_20260719", "全量快照, 可一键还原"),
    ("Top-N 漂移", "Top-3:13 / Top-4:22 / Top-5:953 / Top-6:128", "保留历史原貌, 归一口径=取每位置前5"),
    ("置信度空壳", "84.2% (941/1117) 为空{}, 且 r≈0", "非数据损坏, 属信号缺失, 见'置信度分析'"),
]
r = 6
for a, b, c in rows:
    put(ws, r, 1, a, bold=True)
    put(ws, r, 2, b)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    put(ws, r, 3, c, align=WRAP)
    r += 1

# =====================================================================
# Sheet 2: 问题清单 (带公式比率)
# =====================================================================
ws2 = wb.create_sheet("问题清单")
ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 50

put(ws2, 1, 1, "问题清单与影响面", fill=C_TITLE, bold=True, color=WHITE)
ws2.merge_cells("A1:E1"); ws2.cell(1,1).alignment=CEN; ws2.row_dimensions[1].height=24

hdr = ["#", "问题类型", "记录数", "占比(公式)", "影响说明"]
for i, h in enumerate(hdr, 1):
    put(ws2, 2, i, h)
style_header(ws2, 2, 5)

# 数据 (行3-6); 总记录数在第7行作为分母
data = [
    (1, "嵌套旧格式 (BAD_POS_STRUCT)", 11, "id 10-20: {wan:{numbers,confidence,reason}} 嵌套, 通用工具不可解析"),
    (2, "全空预测记录", 1, "id 29: 五位置全空, 无任何预测号码"),
    (3, "Top-N 版本漂移", 163, "Top-3/4/6 共163条与Top-5混存, 覆盖不可比 (历史原貌)"),
    (4, "置信度空壳", 941, "confidence_scores 为空{}, 且整体与命中相关系数≈0"),
]
r = 3
for n, t, cnt, desc in data:
    put(ws2, r, 1, n, align=CEN)
    put(ws2, r, 2, t, bold=True)
    put(ws2, r, 3, cnt, fill=C_INPUT, align=CEN)   # 硬编码输入(蓝)
    # 占比公式: =C3/$C$8
    f = put(ws2, r, 4, "=C%d/$C$8" % r, align=CEN)
    f.number_format = "0.00%"
    f.font = Font(color=C_BLACK)  # 黑=公式
    put(ws2, r, 5, desc, align=WRAP)
    r += 1
put(ws2, 7, 2, "合计/分母", bold=True)
put(ws2, 7, 3, "=SUM(C3:C6)", align=CEN)  # 公式
put(ws2, 8, 2, "总记录数(基线)", bold=True)
put(ws2, 8, 3, 1117, fill=C_INPUT, align=CEN)   # 硬编码
put(ws2, 8, 4, "分母", align=CEN)
put(ws2, 9, 2, "红线硬伤(问题1+2)", bold=True)
put(ws2, 9, 3, "=C3+C4", align=CEN)
put(ws2, 9, 5, "这两项属结构损坏, 必须修复", align=WRAP)

# =====================================================================
# Sheet 3: 修复动作与前后对比
# =====================================================================
ws3 = wb.create_sheet("修复前后对比")
ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 36
hdr = ["指标", "修复前", "修复后", "说明"]
for i, h in enumerate(hdr, 1):
    put(ws3, 1, i, h)
style_header(ws3, 1, 4)
cmp = [
    ("嵌套旧格式记录数", 11, 0, "11条 id10-20 已转扁平", C_BAD, C_OK),
    ("全空记录数", 1, 1, "id29 隔离为 failed(合法空), 非损坏", C_BAD, C_OK),
    ("红线硬伤总数", 12, 0, "0 违规 = PASS", C_BAD, C_OK),
    ("扁平可读记录数", 1105, 1117, "全部工具可读", None, None),
    ("置信度已恢复记录", 0, 11, "从旧confidence数组恢复 11条", None, None),
    ("Top-N 漂移", "3/4/5/6混存", "3/4/5/6混存", "保留历史, 口径=前5", None, None),
]
r = 2
for name, b, a, note, fb, fa in cmp:
    put(ws3, r, 1, name, bold=True)
    cb = put(ws3, r, 2, b, align=CEN)
    ca = put(ws3, r, 3, a, align=CEN)
    if fb: cb.fill = PatternFill("solid", fgColor=fb)
    if fa: ca.fill = PatternFill("solid", fgColor=fa)
    put(ws3, r, 4, note, align=WRAP)
    r += 1

# =====================================================================
# Sheet 4: Top-N 漂移分析
# =====================================================================
ws4 = wb.create_sheet("TopN漂移")
ws4.column_dimensions["A"].width = 16
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 14
ws4.column_dimensions["D"].width = 50
put(ws4, 1, 1, "Top-N 长度签名分布 (wan,qian,bai,shi,ge 等长)", fill=C_TITLE, bold=True, color=WHITE)
ws4.merge_cells("A1:D1"); ws4.cell(1,1).alignment=CEN
hdr = ["签名", "记录数", "占比", "说明"]
for i, h in enumerate(hdr, 1):
    put(ws4, 2, i, h)
style_header(ws4, 2, 4)
topn = [("(5,5,5,5,5)", 953, "当前生产标准 position_top_n=5"),
        ("(6,6,6,6,6)", 128, "旧版 Top-6, 覆盖比5宽"),
        ("(4,4,4,4,4)", 22, "旧版 Top-4 (含转换的嵌套记录)"),
        ("(3,3,3,3,3)", 13, "旧版 Top-3 (含转换的嵌套记录)"),
        ("(0,0,0,0,0)", 1, "已隔离的空记录 id29")]
r = 3
for sig, cnt, note in topn:
    put(ws4, r, 1, sig, align=CEN)
    put(ws4, r, 2, cnt, fill=C_INPUT, align=CEN)
    f = put(ws4, r, 3, "=B%d/$B$8" % r, align=CEN); f.number_format="0.00%"
    put(ws4, r, 4, note, align=WRAP)
    r += 1
put(ws4, 7, 1, "合计", bold=True); put(ws4, 7, 2, "=SUM(B3:B7)", align=CEN)
put(ws4, 8, 1, "总记录", bold=True); put(ws4, 8, 2, 1117, fill=C_INPUT, align=CEN)
put(ws4, 10, 1, "归一规则 (黄底=重要假设)", fill=C_ASSUME, bold=True)
ws4.merge_cells("A10:D10")
put(ws4, 11, 1, "跨版本命中率对比时, 统一取每位置前 5 个号码为口径 (Top-5 子集)",
    fill=C_ASSUME, align=WRAP)
ws4.merge_cells("A11:D11")
put(ws4, 12, 1, "不截断/不补位历史记录: 改历史会扭曲统计且不可逆, 故仅统一格式+记录口径",
    fill=C_ASSUME, align=WRAP)
ws4.merge_cells("A12:D12")

# =====================================================================
# Sheet 5: 置信度分析 (r=0.000)
# =====================================================================
ws5 = wb.create_sheet("置信度分析")
ws5.column_dimensions["A"].width = 28
ws5.column_dimensions["B"].width = 14
ws5.column_dimensions["C"].width = 14
ws5.column_dimensions["D"].width = 48
put(ws5, 1, 1, "confidence_scores 结构分布与信号有效性", fill=C_TITLE, bold=True, color=WHITE)
ws5.merge_cells("A1:D1"); ws5.cell(1,1).alignment=CEN
hdr = ["结构", "记录数", "占比", "说明"]
for i, h in enumerate(hdr, 1):
    put(ws5, 2, i, h)
style_header(ws5, 2, 4)
conf = [("dict(0 keys) 空", 941, "空壳, 无置信度"),
        ("list(10)", 120, "10个组合置信度"),
        ("dict(5 keys)", 49, "5位置各一置信度"),
        ("list(5)", 7, "5个组合置信度")]
r = 3
for st, cnt, note in conf:
    put(ws5, r, 1, st, bold=True)
    put(ws5, r, 2, cnt, fill=C_INPUT, align=CEN)
    f = put(ws5, r, 3, "=B%d/$B$8" % r, align=CEN); f.number_format="0.00%"
    put(ws5, r, 4, note, align=WRAP)
    r += 1
put(ws5, 7, 1, "合计", bold=True); put(ws5, 7, 2, "=SUM(B3:B6)", align=CEN)
put(ws5, 8, 1, "总记录", bold=True); put(ws5, 8, 2, 1117, fill=C_INPUT, align=CEN)
put(ws5, 10, 1, "关键结论 (r=0.000)", fill=C_ASSUME, bold=True); ws5.merge_cells("A10:D10")
notes = [
    "置信度字段 84.2% 为空壳, 且即便有值的记录, 其置信度与实际命中的相关系数≈0 (参见 v3.14 审计报告)。",
    "因此 confidence 不能作为有效预测信号, 也不应作为'可信度'展示给用户。",
    "建议: 产品层面要么停显 confidence, 要么仅展示已验证可用的统计量(如历史命中率), 避免误导。",
]
rr = 11
for n in notes:
    put(ws5, rr, 1, "· " + n, align=WRAP); ws5.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=4)
    rr += 1

# =====================================================================
# Sheet 6: 所以呢 (建议)
# =====================================================================
ws6 = wb.create_sheet("所以呢")
ws6.column_dimensions["A"].width = 6
ws6.column_dimensions["B"].width = 40
ws6.column_dimensions["C"].width = 16
ws6.column_dimensions["D"].width = 44
put(ws6, 1, 1, "所以呢 · 可执行建议", fill=C_TITLE, bold=True, color=WHITE)
ws6.merge_cells("A1:D1"); ws6.cell(1,1).alignment=CEN
hdr = ["#", "建议", "优先级", "下一步动作"]
for i, h in enumerate(hdr, 1):
    put(ws6, 2, i, h)
style_header(ws6, 2, 4)
recs = [
    (1, "数据红线已修复, 全表现可被任意工具解析", "已完成", "保留备份表30天后清理"),
    (2, "Top-N 漂移按'前5口径'归一分析", "P1", "在统计脚本中统一取每位置前5"),
    (3, "置信度字段停止作为有效信号展示", "P1", "隐藏 confidence 或改显历史命中率"),
    (4, "嵌套格式兼容代码已休眠, 可保留作防御", "P3", "不动 database.py 兼容层"),
    (5, "后续写入统一走 insert_prediction_record", "P2", "防止再出现嵌套/空记录"),
]
r = 3
for n, rec, pr, act in recs:
    put(ws6, r, 1, n, align=CEN)
    put(ws6, r, 2, rec, align=WRAP)
    pfill = C_OK if pr == "已完成" else (C_ASSUME if pr in ("P1","P2") else None)
    put(ws6, r, 3, pr, align=CEN, fill=pfill)
    put(ws6, r, 4, act, align=WRAP)
    r += 1

# 冻结首行 & 保存
for s in wb.worksheets:
    s.freeze_panes = "A2"

out = "reports/diagnostic/排列5_数据质量红线修复报告.xlsx"
wb.save(out)
print("已生成:", out)
