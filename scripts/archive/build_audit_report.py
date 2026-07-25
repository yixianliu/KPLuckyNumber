# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
build_audit_report.py — 生成《排列5 v3.14 算法/策略/流程审计与命中率报告》
(2026-07-19 审计任务)

数据来源(可追溯):
  - 训练窗口寻优: opt_window_study.py (2026-07-19, anaconda3, DB localhost/lucky_number, 1010期历史,
    最近120期 walk-forward, AI关闭, 截断传入 predict 的历史长度实现固定窗口)
  - 稳健性3折: 内联脚本 (last180期分3折各60期, OLD=lookback=None/recency关 vs NEW=lookback=60/recency开)
  - 配置改动: modules/predictor.py DEFAULT_CONFIG + _algo_frequency_weighted (v3.14审计)
所有数值为实验实测, 非硬编码假设。
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- 颜色规范 (data-analysis-expert) ----
C_INPUT   = "FFCCE5FF"  # 蓝: 硬编码输入/配置值
C_ASSUM   = "FFFFFF99"  # 黄: 关键假设项
C_COMPUTE = "FFFFFFFF"  # 黑字(计算值, 白底)
C_HDR     = "FF4472C4"  # 表头蓝
C_GOOD    = "FFC6EFCE"  # 绿: 达标
C_WARN    = "FFFFC7CE"  # 红: 警示
C_NTRL    = "FFF2F2F2"  # 中性灰

thin = Side(style="thin", color="FFBBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FONT = Font(bold=True, color="FFFFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="FF1F4E78")
SUB_FONT = Font(bold=True, size=11, color="FF1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

wb = Workbook()

def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=C_HDR)
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = border

def put_table(ws, start_row, headers, rows, colw=None, fill_rule=None):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.border = border
            cell.alignment = WRAP if isinstance(val, str) and len(str(val)) > 12 else CENTER
            if fill_rule:
                f = fill_rule(j, val, row)
                if f:
                    cell.fill = PatternFill("solid", fgColor=f)
        r += 1
    if colw:
        for j, w in enumerate(colw, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return r

# =====================================================================
# Sheet 1: 审计概览
# =====================================================================
ws = wb.active
ws.title = "审计概览"
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 90
ws['A1'] = "排列5 AI智能分析系统 — v3.14 算法/策略/流程审计与命中率报告"
ws['A1'].font = TITLE_FONT
ws.merge_cells('A1:B1')

overview = [
    ("审计日期", "2026-07-19 (GMT+8)"),
    ("审计范围", "核心算法逻辑 / 策略与走势吻合度 / 预测流程优化 / 命中率趋势与归因 / 最优训练窗口"),
    ("数据来源", "DB localhost:3306/lucky_number.p5_history_data (1010期, issue 2023234→2026189); AI关闭纯统计模型"),
    ("当前版本", "v3.14 — 双信号自适应架构(默认关闭); 频率权重0.54/遗漏0.34/贝叶斯0.10/次要≤0.01"),
    ("", ""),
    ("★ 核心结论1", "训练窗口对命中率无统计显著影响: 9个窗口(30~ALL) walk-forward 的 Top-1 全部在 8.5%~12.0% 间, "
                     "均值≈10%(随机基线)。窗口=40 单点12.0%是3折验证中某折(最近60期13.67%)的偶然, 非稳健优势。"),
    ("★ 核心结论2", "v3.12宣称的'11.25~11.75%实质性提升'真实但非稳健: 集中于近期有利窗口, 跨全历史回归至≈10%。属抽样噪声, 非可持续信号。"),
    ("★ 核心结论3", "排列5为公平摇号, 无算法能从历史提取稳定超越随机的预测力。产品价值在'分析质量/可解释性', 而非'命中率'。"),
    ("★ 核心结论4", "审计发现并修复1处死配置: frequency.recency_weight=True 此前从未被读取(等权计数); 已落地为真实近期指数加权实现(默认关, 数据中性偏负)。"),
    ("★ 最优窗口决策", "lookback_periods: None→60 (用户授权自行决定)。理由: 非为提升命中率(无显著影响), "
                       "而为防止历史无限增长后频率经验分布退化至完全均匀→主信号消失。60期为近期响应性与稳定性的折中。"),
    ("", ""),
    ("所以呢？(建议)", "① 接受当前v3.14静态默认(≈随机)为已知最优, 停止在'命中率'维度过度调参; "
                      "② 把产品重心转向分析可读性/导出/爬取增量等用户侧价值; "
                      "③ 若坚持探索, 唯一有意义方向是'条件信号'(特定号码形态下切换算法), 但需先定义可验证假设; "
                      "④ 修复历史数据质量(早期诊断发现 predicted_numbers Top-N 版本漂移/11条损坏), 使验证证据可信。"),
]
r = 3
for k, v in overview:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, color="FF1F4E78")
    ws.cell(row=r, column=1).alignment = WRAP
    c = ws.cell(row=r, column=2, value=v)
    c.alignment = WRAP
    if k.startswith("★"):
        c.fill = PatternFill("solid", fgColor=C_ASSUM)
    r += 1
for rr in range(3, r):
    ws.row_dimensions[rr].height = 42

# =====================================================================
# Sheet 2: 训练窗口寻优
# =====================================================================
ws2 = wb.create_sheet("训练窗口寻优")
ws2['A1'] = "训练窗口寻优 (walk-forward, 最近120期, AI关闭, 截断传入历史长度实现固定窗口)"
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:G1')
ws2['A2'] = "数据来源: opt_window_study.py (2026-07-19, anaconda3, DB localhost/lucky_number, 1010期历史)"
ws2['A2'].font = Font(italic=True, size=9, color="FF808080")
ws2.merge_cells('A2:G2')

headers = ["训练窗口(期)", "Top-1%", "Top-3%", "Top-5%", "Top-6%", "avg_mc", "mc_std(稳定性)"]
# 实测 (来自 window_study.json stdout)
rows = [
    [30, 11.0, 32.83, 50.17, 59.17, 0.550, 0.693],
    [40, 12.0, 32.33, 50.50, 59.17, 0.600, 0.757],
    [50, 11.0, 29.00, 47.17, 60.50, 0.550, 0.740],
    [60, 9.67, 27.50, 47.50, 58.83, 0.483, 0.683],
    [80, 10.33, 25.83, 50.17, 59.50, 0.517, 0.671],
    [100, 10.17, 26.83, 45.50, 56.50, 0.508, 0.730],
    [120, 8.50, 27.17, 49.00, 58.00, 0.425, 0.691],
    [150, 8.83, 28.00, 47.50, 59.17, 0.442, 0.693],
    [200, 9.67, 26.33, 47.33, 58.00, 0.483, 0.695],
    ["ALL(全量)", 11.83, 32.83, 51.00, 60.50, 0.592, 0.639],
]
rand = [0, 10.0, 30.0, 50.0, 60.0, 3.0, 0.0]
def wfill(j, val, row):
    if j == 1 and val == 40:
        return C_ASSUM  # 标注单点最优
    if j == 1 and val == "ALL(全量)":
        return C_INPUT
    if j == 2:  # Top-1 着色
        if isinstance(val, (int, float)) and val >= 11.5:
            return C_GOOD
        if isinstance(val, (int, float)) and val < 9.5:
            return C_WARN
    return None
endr = put_table(ws2, 4, headers, rows, colw=[16, 10, 10, 10, 10, 10, 16], fill_rule=wfill)
# 随机基线行
ws2.cell(row=endr, column=1, value="随机基线").font = Font(bold=True, color="FF808080")
for j, v in enumerate(rand, 1):
    cell = ws2.cell(row=endr, column=j, value=v)
    cell.border = border
    cell.alignment = CENTER
    cell.font = Font(italic=True, color="FF808080")
    cell.fill = PatternFill("solid", fgColor=C_NTRL)
ws2.cell(row=endr+2, column=1, value="解读: 各窗口Top-1在8.5%~12.0%剧烈摆动, 无单调趋势; 中窗口(60~150)反而最差。"
        "窗口选择不是命中率杠杆。决策: lookback=60(防长期退化), 非为提命中率。").alignment = WRAP
ws2.merge_cells(start_row=endr+2, start_column=1, end_row=endr+2, end_column=7)
ws2.row_dimensions[endr+2].height = 30

# =====================================================================
# Sheet 3: 稳健性验证 (3折)
# =====================================================================
ws3 = wb.create_sheet("稳健性验证")
ws3['A1'] = "稳健性 3折交叉验证 (最近180期, 每折60期)"
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:F1')
ws3['A2'] = "数据来源: 内联脚本 (2026-07-19) — OLD=lookback=None/recency关; NEW=lookback=60/recency开"
ws3['A2'].font = Font(italic=True, size=9, color="FF808080")
ws3.merge_cells('A2:F2')
h3 = ["配置", "折1 T1/mc", "折2 T1/mc", "折3 T1/mc", "均值T1%", "结论"]
r3 = [
    ["OLD (lookback=None, recency关)", "6.33/0.317", "11.33/0.567", "12.33/0.617", 10.0, "≈随机, 折间摆动大"],
    ["NEW (lookback=60, recency开)", "7.67/0.383", "8.33/0.417", "13.0/0.65", 9.67, "中性偏负, 未提升"],
    ["当前默认 (lookback=60, recency关)", "—", "—", "—", "≈10.0*", "*介于OLD与NEW之间, 推断≈随机"],
]
endr3 = put_table(ws3, 4, h3, r3, colw=[32, 14, 14, 14, 12, 26])
ws3.cell(row=endr3+1, column=1, value="结论: 任何窗口/近期加权配置, 3折均值均≈10%(随机)。单折6%~14%的摆动远超配置间差异→配置非命中率杠杆。").alignment = WRAP
ws3.merge_cells(start_row=endr3+1, start_column=1, end_row=endr3+1, end_column=6)
ws3.row_dimensions[endr3+1].height = 30

# =====================================================================
# Sheet 4: 核心算法审计
# =====================================================================
ws4 = wb.create_sheet("核心算法审计")
ws4['A1'] = "7算法融合逻辑审计"
ws4['A1'].font = TITLE_FONT
ws4.merge_cells('A1:F1')
h4 = ["算法", "权重", "核心逻辑", "理论合理性", "审计发现", "处置"]
r4 = [
    ["frequency_weighted", "0.54", "近N期各位置号码经验频率(拉普拉斯平滑MLE)", "高(对i.i.d.随机是理论最优主信号)",
     "recency_weight=True此前是死配置(等权计数); lookback=None全量→历史增长退化至均匀", "已落地近期指数加权实现; lookback=60"],
    ["omission_regression", "0.34", "冷号回补: 遗漏越大概率越高(exp(β·o))", "中(赌徒谬误反向信号, 与频率互补)",
     "逻辑正确; max_omission_cap=50防exp爆炸合理", "保持"],
    ["bayesian_inference", "0.10", "经验加权(非严格贝叶斯)", "低(KNOWN_ISSUES: 似然无界累乘失真)",
     "文档自述'经验加权而非严格贝叶斯'", "保持(标记为示意)"],
    ["trend_momentum", "0.01", "近30期线性回归斜率+高斯衰减", "低(随机游走斜率无预测力)", "窗口30固定, 未与频率窗口协同", "保持(学习通道)"],
    ["markov_transition", "0.005", "一阶转移概率", "低(相邻期独立)", "噪声源", "保持(学习通道)"],
    ["pattern_continuation", "0.003", "近7期形态延续", "低", "窗口7固定", "保持(学习通道)"],
    ["feature_engineering", "0.002", "多维特征(含贝叶斯特征)", "低", "v3.12补齐, 此前缺失", "保持(学习通道)"],
]
put_table(ws4, 3, h4, r4, colw=[20, 8, 34, 30, 36, 20])
for rr in range(4, 4+len(r4)):
    ws4.row_dimensions[rr].height = 40

# =====================================================================
# Sheet 5: 流程优化项
# =====================================================================
ws5 = wb.create_sheet("流程优化项")
ws5['A1'] = "预测流程可优化环节 (审计发现)"
ws5['A1'].font = TITLE_FONT
ws5.merge_cells('A1:D1')
h5 = ["问题", "严重度", "说明", "处理"]
r5 = [
    ["frequency.recency_weight 死配置", "中", "配置声明True但频率算法从不读取, 文档宣称'近期加权'未实现, 误导", "已落地真实近期指数加权(默认关, 数据中性)"],
    ["frequency.lookback=None 全量", "中", "随历史增长→经验频率趋近均匀→主信号消失(长期退化)", "改为60期有界窗口(防退化, 命中率无显著影响)"],
    ["各算法窗口独立硬编码", "低", "趋势30/形态7/贝叶斯60/频率原None, 未统一协同优化", "记录为待办(影响小, 暂不重构)"],
    ["自适应权重(双信号)默认关", "低", "v3.14实验: Top-1信号CV=0.47但30期样本EWMA噪声>信号, 自适应略输基线", "保持关闭, 待真生产500+期再验"],
    ["历史数据质量红线", "高", "早期诊断: predicted_numbers Top-N版本漂移(948条Top-5/17Top-4/8Top-6/11损坏); confidence r=0.000", "建议清数据并重生成验证证据(未授权修改)"],
    ["predict --model old/optimized 等价", "低", "KNOWN_ISSUES: use_optimized形参被忽略, 对比模式无差异", "记录, 暂不处理"],
]
def sev_fill(j, val, row):
    if j == 2:
        if val == "高": return C_WARN
        if val == "中": return C_ASSUM
    return None
put_table(ws5, 3, h5, r5, colw=[28, 10, 56, 40], fill_rule=sev_fill)
for rr in range(4, 4+len(r5)):
    ws5.row_dimensions[rr].height = 38

# =====================================================================
# Sheet 6: 命中率归因
# =====================================================================
ws6 = wb.create_sheet("命中率归因")
ws6['A1'] = "命中率变化趋势与归因分析"
ws6['A1'].font = TITLE_FONT
ws6.merge_cells('A1:C1')
h6 = ["维度", "观察", "归因"]
r6 = [
    ["Top-1 区间", "各窗口8.5%~12.0%; 3折6%~14%", "等于/略超随机(10%); 摆动为抽样噪声, 非信号"],
    ["v3.12 宣称提升", "11.25~11.75% (近80期walk-forward)", "集中于近期有利窗口; 跨全历史回归≈10% → 非稳健"],
    ["窗口效应", "中窗口(60~150)最差, 近期40/全量并列最佳", "倒U型: 近期40≈全量(均被近期经验主导); 中窗口稀释"],
    ["算法贡献", "频率独跑Top-1≈11.75%(=融合), 其余算法为噪声", "频率是绝对主力; 次要算法稀释信号(故压至≤0.01)"],
    ["自适应影响", "开自适应→Top-1降至9.5%(EWMA学歪频率0.54→0.42)", "零信号时EWMA放大噪声→已默认关闭"],
    ["根本上限", "公平摇号, 期望Top-1=10%", "任何统计量都无法稳定突破随机基线"],
]
put_table(ws6, 3, h6, r6, colw=[18, 48, 50])
for rr in range(4, 4+len(r6)):
    ws6.row_dimensions[rr].height = 34

# =====================================================================
# Sheet 7: 建议与下一步
# =====================================================================
ws7 = wb.create_sheet("建议与下一步")
ws7['A1'] = "所以呢？ —— 可执行建议"
ws7['A1'].font = TITLE_FONT
ws7.merge_cells('A1:B1')
h7 = ["优先级", "建议"]
r7 = [
    ["P0", "接受v3.14静态默认(≈随机)为已知最优; 停止在'命中率'维度过度调参(已证无杠杆)。"],
    ["P0", "修复历史数据质量红线: 统一predicted_numbers口径、修复11条损坏记录, 使验证证据可信(需授权)。"],
    ["P1", "产品重心转向用户侧价值: 分析可读性、Excel/PDF导出、爬取增量、AI报告重写。"],
    ["P2", "若坚持探索预测: 唯一有意义方向='条件信号'(特定号码形态/冷热态下切换算法); 需先定义可验证假设再实验。"],
    ["P2", "保留双信号自适应架构(已落地), 待真生产500+期干净验证数据后, 再开enable_adaptive_weights实验。"],
    ["P3", "统一各算法窗口为可配置中心参数(消除趋势30/形态7/贝叶斯60硬编码), 降低维护认知负担。"],
]
put_table(ws7, 3, h7, r7, colw=[12, 95])
for rr in range(4, 4+len(r7)):
    ws7.row_dimensions[rr].height = 34

wb.save("reports/diagnostic/排列5_v3.14算法策略流程审计报告.xlsx")
print("saved reports/diagnostic/排列5_v3.14算法策略流程审计报告.xlsx")
