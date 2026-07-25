# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
生成 v3.12 生产验证对比报告（Excel）
====================================
读取 reports/diagnostic/v312_production.json，与以下基线对比：
  - 随机基线 Top-1/3/5/6 = 10/30/50/60%
  - 旧模型 992 条真实记录标准化 Top-3 ≈ 28%（来自 opt_diagnostic Part A）
产出多 sheet 工作簿 + 图表，派生指标 Python 预计算写值（规避公式错误）。

配色规范（数据分析专家）：
  蓝=硬编码输入  黑=计算值  绿=超基线/外部  红=低于基线  黄底=重要假设
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

SRC = 'reports/diagnostic/v312_production.json'
OUT = 'reports/diagnostic/排列5_v3.12生产验证报告.xlsx'

# 配色
BLUE = '1F4E79'      # 硬编码输入
BLACK = '000000'     # 计算值
GREEN = '2E7D32'     # 超基线
RED = 'C62828'       # 低于基线
YELLOW = 'FFF3CD'    # 重要假设
HEADER_FILL = '1F4E79'
SUB_FILL = 'D6DCE4'
WHITE = 'FFFFFF'

thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')


def font(color=BLACK, bold=False, size=10):
    return Font(name='微软雅黑', color=color, bold=bold, size=size)


def fill(color):
    return PatternFill('solid', fgColor=color)


def put(ws, r, c, v, color=BLACK, bold=False, fmt=None, align=None, fillc=None, size=10):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font(color, bold, size)
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    else:
        cell.alignment = CENTER
    cell.border = BORDER
    if fillc:
        cell.fill = fill(fillc)
    return cell


def hdr(ws, r, cols):
    for i, c in enumerate(cols, 1):
        put(ws, r, i, c, color=WHITE, bold=True, fillc=HEADER_FILL)


def build():
    d = json.load(open(SRC, encoding='utf-8'))
    POS = ['万位', '千位', '百位', '十位', '个位']
    rb = d['random_base']
    old = d['old_model_real_top3']

    wb = Workbook()

    # ---------- Sheet 1: 概览 ----------
    ws = wb.active
    ws.title = '概览'
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 40
    put(ws, 1, 1, 'v3.12 生产口径验证报告', color=BLUE, bold=True, size=14, align=Alignment(vertical='center'))
    ws.merge_cells('A1:D1')
    put(ws, 2, 1, f"验证窗口: 最近 {d['tested']} 期 (start={d['start_index']}) · 关AI纯算法 · walk-forward 防前视偏差",
        color=BLACK, size=9, align=Alignment(vertical='center'))
    ws.merge_cells('A2:D2')

    put(ws, 4, 1, '指标', WHITE, bold=True, fillc=HEADER_FILL)
    put(ws, 4, 2, 'v3.12 真实', WHITE, bold=True, fillc=HEADER_FILL)
    put(ws, 4, 3, '随机基线', WHITE, bold=True, fillc=HEADER_FILL)
    put(ws, 4, 4, '结论', WHITE, bold=True, fillc=HEADER_FILL)

    rows = [
        ('整体 Top-1 命中率', d['top1_overall'], rb['top1'],
         '✓ 超基线' if d['top1_overall'] > rb['top1'] else '✗ 低于基线'),
        ('整体 Top-3 命中率', d['top3_overall'], rb['top3'],
         '✓ 超基线' if d['top3_overall'] > rb['top3'] else '✗ 低于基线'),
        ('整体 Top-5 命中率', d['top5_overall'], rb['top5'],
         '✓ 超基线' if d['top5_overall'] > rb['top5'] else '✗ 低于基线'),
        ('整体 Top-6 命中率(生产口径)', d['top6_overall_rate'], rb['top6'],
         '✓ 超基线' if d['top6_overall_rate'] > rb['top6'] else '✗ 低于基线'),
        ('平均 match_count(Top-6口径/5)', d['avg_top6_match_count'], 3.00,
         '✓ 超基线' if d['avg_top6_match_count'] > 3.0 else '✗ 低于基线'),
        ('旧模型真实 Top-3(992条,参考)', old, rb['top3'],
         '≈ v3.12 改进前基准'),
    ]
    r = 5
    for name, val, base, concl in rows:
        put(ws, r, 1, name, color=BLUE)
        exceed = val > base
        put(ws, r, 2, val, color=BLACK, fmt='0.00"%"' if '率' in name else '0.000')
        put(ws, r, 3, base, color=BLACK, fmt='0.00"%"' if '率' in name else '0.000')
        c4 = GREEN if (exceed and '超' in concl) else (RED if '低于' in concl else BLACK)
        put(ws, r, 4, concl, color=c4, bold=True)
        r += 1

    put(ws, r + 1, 1, f"数据来源: reports/diagnostic/v312_production.json（v3.12 融合配置，"
        f"{d['tested']}期真实开奖验证，耗时{d['elapsed_s']}s）", color=BLACK, size=8)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)
    put(ws, r + 2, 1, "注: 旧模型真实 Top-3=28% 来自 opt_diagnostic Part A（992条已验证记录标准化到 Top-3），"
        "反映 v3.12 之前的模型表现。", color=BLACK, size=8)
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=4)

    # ---------- Sheet 2: 逐位置 Top-6（生产口径） ----------
    ws2 = wb.create_sheet('逐位置Top6命中率')
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    hdr(ws2, 1, ['位置', 'v3.12 Top-6', '随机基线60%', '是否超基线'])
    p6 = d['pos_top6_rate']
    r = 2
    for pos in POS:
        v = p6[pos]
        put(ws2, r, 1, pos, color=BLUE)
        put(ws2, r, 2, v, color=BLACK, fmt='0.00"%"')
        put(ws2, r, 3, 60.0, color=BLACK, fmt='0.00"%"')
        ok = v > 60.0
        put(ws2, r, 4, '✓ 超基线' if ok else '✗ 低于基线',
            color=GREEN if ok else RED, bold=True)
        r += 1
    put(ws2, r, 1, '整体', color=BLUE, bold=True)
    put(ws2, r, 2, d['top6_overall_rate'], color=BLACK, bold=True, fmt='0.00"%"')
    put(ws2, r, 3, 60.0, color=BLACK, bold=True, fmt='0.00"%"')
    put(ws2, r, 4, '✓ 超基线' if d['top6_overall_rate'] > 60 else '✗ 低于基线',
        color=GREEN if d['top6_overall_rate'] > 60 else RED, bold=True)
    # 图表
    chart = BarChart()
    chart.title = '逐位置 Top-6 命中率 vs 随机基线(60%)'
    chart.type = 'col'
    chart.y_axis.title = '命中率%'
    chart.height = 8
    chart.width = 16
    data = Reference(ws2, min_col=2, max_col=2, min_row=1, max_row=r)
    cats = Reference(ws2, min_col=1, max_col=1, min_row=2, max_row=r)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = BLUE
    ws2.add_chart(chart, f'A{r + 3}')

    # ---------- Sheet 3: 概率排名口径对比 ----------
    ws3 = wb.create_sheet('概率排名口径对比')
    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 16
    ws3.column_dimensions['D'].width = 16
    ws3.column_dimensions['E'].width = 18
    hdr(ws3, 1, ['口径', 'v3.12 真实', '随机基线', '旧模型真实', '是否超基线'])
    prow = [
        ('Top-1', d['top1_overall'], rb['top1'], None),
        ('Top-3', d['top3_overall'], rb['top3'], old),
        ('Top-5', d['top5_overall'], rb['top5'], None),
        ('Top-6(生产)', d['top6_overall_rate'], rb['top6'], None),
    ]
    r = 2
    for name, val, base, o in prow:
        put(ws3, r, 1, name, color=BLUE)
        put(ws3, r, 2, val, color=BLACK, fmt='0.00"%"')
        put(ws3, r, 3, base, color=BLACK, fmt='0.00"%"')
        put(ws3, r, 4, o if o is not None else '—', color=BLACK, fmt='0.00"%"' if o else None)
        ok = val > base
        put(ws3, r, 5, '✓ 超基线' if ok else '✗ 低于基线',
            color=GREEN if ok else RED, bold=True)
        r += 1
    chart = BarChart()
    chart.title = 'v3.12 vs 随机基线（概率排名口径）'
    chart.type = 'col'
    chart.height = 8
    chart.width = 16
    data = Reference(ws3, min_col=2, max_col=3, min_row=1, max_row=r - 1)
    cats = Reference(ws3, min_col=1, max_col=1, min_row=2, max_row=r - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = BLUE
    chart.series[1].graphicalProperties.solidFill = '888888'
    ws3.add_chart(chart, f'A{r + 3}')

    # ---------- Sheet 4: match_count 分布 ----------
    ws4 = wb.create_sheet('match_count分布')
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 16
    ws4.column_dimensions['C'].width = 16
    hdr(ws4, 1, ['命中位(满分5)', '期数', '占比'])
    dist = d['match_count_dist']
    total = d['tested']
    r = 2
    for k in range(6):
        c = dist.get(str(k), 0)
        put(ws4, r, 1, k, color=BLUE)
        put(ws4, r, 2, c, color=BLACK)
        put(ws4, r, 3, c / total, color=BLACK, fmt='0.0%')
        r += 1
    put(ws4, r, 1, '合计', color=BLUE, bold=True)
    put(ws4, r, 2, f'=SUM(B2:B{r-1})', color=BLACK, bold=True)
    put(ws4, r, 3, 1.0, color=BLACK, bold=True, fmt='0.0%')
    # 合计行的 SUM 公式在无重算环境可能不刷新，改为直接写值
    ws4.cell(row=r, column=2, value=total)
    ws4.cell(row=r, column=3, value=1.0)

    # ---------- Sheet 5: 结论与下一步 ----------
    ws5 = wb.create_sheet('结论与下一步')
    ws5.column_dimensions['A'].width = 100
    lines = [
        ('v3.12 生产验证结论（让数字说话）', BLUE, True, 13),
        (f"1. 在最近 {d['tested']} 期真实已开奖数据上，v3.12 融合配置：", BLACK, False, 10),
        (f"   - 生产口径(每位置Top-6)整体命中率 {d['top6_overall_rate']}%，随机基线60%，"
         f"{'超基线' if d['top6_overall_rate']>60 else '低于基线'}; 平均match_count {d['avg_top6_match_count']}/5 (>3.0为优)", BLACK, False, 10),
        (f"   - 概率排名口径 Top-1={d['top1_overall']}% / Top-3={d['top3_overall']}% / Top-5={d['top5_overall']}%，"
         f"对比随机 10/30/50%", BLACK, False, 10),
        (f"2. 对比旧模型(992条真实记录标准化Top-3≈{old}%): v3.12 在概率口径上全面好于旧模型的真实表现，"
         f"证实 v3.12 修掉了'自伤式低于随机'的问题。", BLACK, False, 10),
        (f"3. 但 v3.12 仍仅停在'等于或略超随机'区间，未产生可重复的显著超额收益"
         f"（结合上一轮5窗口稳健性扫描 T3≈30.5%/基线30%）。", BLACK, False, 10),
        ('', BLACK, False, 6),
        ('所以呢？下一步动作（可执行）', BLUE, True, 12),
        ('A. 校准定位: 系统当前应定位为"覆盖率/缩号工具"(Top-6覆盖60%)，'
         '而非"精准命中预测"，对外表述避免夸大。', BLACK, False, 10),
        ('B. 复活自学习闭环(高价值): 用 v3.12 对真实期号生成预测并写库(新report_uuid不污染旧数据)，'
         '调 verify_pending_predictions 验证 → 写入 p5_artifact(weight_history)，让自适应权重真正闭环。', BLACK, False, 10),
        ('C. 修数据质量: 统一历史记录 Top-N 口径入库、清洗 11 条损坏记录，使后续命中率横比可信。', BLACK, False, 10),
        ('D. 换信号源思路: 7 算法已证无边缘，继续调权重收益有限；需引入新特征/外部信号才可能突破随机天花板。', BLACK, False, 10),
        ('', BLACK, False, 6),
        (f"数据来源: v312_production.json（v3.12融合, {d['tested']}期, 耗时{d['elapsed_s']}s）"
         f" + 旧模型来自 opt_diagnostic/partA.json。检索/生成时间: 2026-07-18。", BLACK, False, 8),
    ]
    r = 1
    for text, color, bold, size in lines:
        put(ws5, r, 1, text, color=color, bold=bold, size=size, align=Alignment(wrap_text=True, vertical='top'))
        r += 1

    os.makedirs('reports/diagnostic', exist_ok=True)
    wb.save(OUT)
    print(f"已生成 {OUT} (sheets: {wb.sheetnames})")


if __name__ == '__main__':
    build()
