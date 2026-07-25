# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""Plan B 报告：防过均匀化 — 权重配置对比 + 应用后验证（预计算写值，无公式）"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

BASE = 'D:/PythonProject/KPLuckyNumber'
plan = json.load(open(f'{BASE}/reports/diagnostic/plan_b.json', encoding='utf-8'))
applied = json.load(open(f'{BASE}/reports/diagnostic/v312_production.json', encoding='utf-8'))

BLUE = '1F4E78'; BLACK = '000000'; GREEN = '2E7D32'; RED = 'C62828'
YELLOW = 'FFF2CC'; GREY = 'F2F2F2'
WHITE = 'FFFFFF'
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
F_BOLD = Font(bold=True, color=WHITE)
F_NORM = Font(color=BLACK)
ALGO_CN = {'frequency_weighted': '频率', 'omission_regression': '遗漏',
           'bayesian_inference': '贝叶斯', 'trend_momentum': '趋势',
           'markov_transition': '马尔可夫', 'pattern_continuation': '形态',
           'feature_engineering': '特征'}


def font(color=BLACK, bold=False):
    return Font(color=color, bold=bold, name='微软雅黑', size=10)


def fill(color):
    return PatternFill('solid', fgColor=color)


def hdr(ws, r, cols, color=BLUE):
    for c, txt in enumerate(cols, 1):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = F_BOLD
        cell.fill = fill(color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def put(ws, r, c, v, color=BLACK, bold=False, fmt=None, align='center', fillc=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font(color, bold)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if fillc:
        cell.fill = fill(fillc)
    return cell


wb = openpyxl.Workbook()

# ===== Sheet 1: 概览 =====
ws = wb.active
ws.title = '概览'
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 78
put(ws, 1, 1, 'Plan B 防过均匀化 — 深度诊断与落地', color=WHITE, bold=True, fillc=BLUE)
ws.merge_cells('A1:B1')
ws.row_dimensions[1].height = 24

rows = [
    ('背景', '复活自学习闭环后，EWMA 把所有算法权重拉向均匀（各算法真实命中率都≈随机0.5），'
            '导致覆盖类指标(Top-5/6)略升，但 Top-1 精准度从 11.5% 跌到 8.33%（低于随机10%）。'),
    ('根因', 'EWMA 学的是「覆盖命中率」(每算法 Top-N 包含率)，而 7 个算法覆盖命中率都≈0.5，'
            '无区分信号。任何非平凡混合都会稀释频率算法主导的 Top-1 精准度。'),
    ('实验', '对最近 120 期真实已开奖 walk-forward 验证 6 种权重配置（关AI、关自适应、自定义权重）：'
            'default / EWMA无约束 / 封顶0.10 / 封顶0.05 / alpha0.15+封顶 / 混合0.1+封顶 / 静态核心。'),
    ('核心发现', '所有 EWMA 配置均把 Top-1 压到 ~8%（低于随机）；只有静态默认保住 Top-1=11.5%。'
            '封顶/调alpha 改善覆盖均匀度但救不了 Top-1；唯一保住 Top-1 的是静态默认。'),
    ('落地配置(G)', 'ewma_blend 0.3→0.1（让静态默认主导、EWMA仅微调）+ minor_max_weight=0.10（次要算法封顶）。'
            '这是实测最佳自适应配置：Top-1 从 8.33% 回升到 9.33%，Top-5/6 保持 52.0%/61.83%（均超随机）。'),
    ('应用后验证', '改 predictor.py 后重跑生产验证（自适应默认开，自动回放120条weight_history）：'
            'Top-1=9.33% / Top-3=30.67% / Top-5=52.0% / Top-6=61.83% / mc=3.092。'),
    ('诚实结论', 'G 配置严格优于无约束 EWMA（Top-1 +1.0pp，覆盖不塌），且防过均匀。'
            '但自适应循环仍略逊于静态默认(11.5%)——因学的指标是覆盖非精准。'),
    ('下一步建议', '① 若 Top-1 优先级最高：可将 ewma_blend 进一步降到 0.05，或直接关自适应用静态默认；'
            '② 根治：把自适应评估指标从「覆盖命中率」改为「Top-1 精准度」，让 EWMA 能主动boost频率算法；'
            '③ 持续积累验证记录(>300期)让 EWMA 更稳健。'),
]
r = 3
for k, v in rows:
    put(ws, r, 1, k, color=BLUE, bold=True, align='center', fillc=GREY)
    put(ws, r, 2, v, align='left')
    ws.row_dimensions[r].height = 46
    r += 1

# ===== Sheet 2: 配置对比 =====
ws = wb.create_sheet('配置对比')
labels = [x['label'] for x in plan['results']]
metrics = {x['label']: x for x in plan['results']}
# 应用后 G（从生产验证json，应用代码跑的，权威）
applied_g = {'label': 'G_应用后(权威)', 'top1_overall': applied['top1_overall'],
             'top3_overall': applied['top3_overall'], 'top5_overall': applied['top5_overall'],
             'top6_overall': applied['top6_overall_rate'], 'avg_match_count': applied['avg_top6_match_count']}
# 顺序展示
order = ['A_default', 'B_ewma_unconstrained', 'C_cap_minor_0.10', 'D_cap_minor_0.05',
         'E_alpha0.15_cap0.10', 'G_blend0.1_cap0.10', 'F_keep_core']
disp = []
for l in order:
    if l in metrics:
        disp.append(metrics[l])
disp.append(applied_g)

cols = ['配置', 'Top-1%', 'Top-3%', 'Top-5%', 'Top-6%', 'match_count', '说明']
hdr(ws, 1, cols)
widths = [22, 9, 9, 9, 9, 12, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
r = 2
notes = {
    'A_default': 'v3.12 静态默认(基准)',
    'B_ewma_unconstrained': '原无约束EWMA(改动前状态)',
    'C_cap_minor_0.10': '封顶0.10',
    'D_cap_minor_0.05': '封顶0.05',
    'E_alpha0.15_cap0.10': 'alpha0.15+封顶(覆盖最好)',
    'G_blend0.1_cap0.10': '混合0.1+封顶(实验最佳)',
    'F_keep_core': '静态核心(忽略EWMA)',
    'G_应用后(权威)': '★落地配置:改后代码重跑',
}
for x in disp:
    is_app = x['label'] == 'G_应用后(权威)'
    fc = YELLOW if is_app else None
    put(ws, r, 1, x['label'], bold=is_app, fillc=fc, align='left')
    put(ws, r, 2, x['top1_overall'], color=(RED if x['top1_overall'] < 10 else BLACK), bold=is_app, fillc=fc)
    put(ws, r, 3, x['top3_overall'], color=(GREEN if x['top3_overall'] >= 30 else BLACK), fillc=fc)
    put(ws, r, 4, x['top5_overall'], color=(GREEN if x['top5_overall'] >= 50 else BLACK), fillc=fc)
    put(ws, r, 5, x['top6_overall'], color=(GREEN if x['top6_overall'] >= 60 else BLACK), fillc=fc)
    put(ws, r, 6, x['avg_match_count'], fillc=fc)
    put(ws, r, 7, notes.get(x['label'], ''), align='left', fillc=fc)
    r += 1
# 随机基线行
put(ws, r, 1, '随机基线', color=GREEN, bold=True)
for c, val in [(2, 10), (3, 30), (4, 50), (5, 60)]:
    put(ws, r, c, val, color=GREEN, bold=True)
put(ws, r, 6, 3.00, color=GREEN, bold=True)
put(ws, r, 7, '理论随机期望', color=GREEN, align='left')
baseline_row = r

# Top-1 柱状图
chart = BarChart()
chart.type = 'col'
chart.title = '各配置 Top-1 命中率(%) 对比'
chart.y_axis.title = 'Top-1 %'
chart.x_axis.title = '配置'
data = Reference(ws, min_col=2, min_row=1, max_row=baseline_row)
cats = Reference(ws, min_col=1, min_row=2, max_row=baseline_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 8
chart.width = 18
ws.add_chart(chart, 'I2')

# ===== Sheet 3: 权重分布 =====
ws = wb.create_sheet('权重分布')
cw = plan['candidates_weights']
algo_order = ['frequency_weighted', 'omission_regression', 'bayesian_inference',
              'trend_momentum', 'markov_transition', 'pattern_continuation', 'feature_engineering']
hdr(ws, 1, ['算法'] + [l for l in order])
ws.column_dimensions['A'].width = 14
for i in range(len(order)):
    ws.column_dimensions[get_column_letter(i + 2)].width = 16
r = 2
for a in algo_order:
    put(ws, r, 1, ALGO_CN[a], bold=True, align='left')
    for c, l in enumerate(order, 2):
        w = cw[l][a]
        fillc = YELLOW if (l == 'G_blend0.1_cap0.10') else None
        put(ws, r, c, round(w, 4), fillc=fillc, fmt='0.000')
    r += 1
put(ws, r, 1, '说明', bold=True, align='left')
put(ws, r, 2, '★黄列=Plan B落地配置(G): 核心(freq/omi/bayes)主导, 次要算法受0.10封顶钳制', align='left')
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(order) + 1)
ws.row_dimensions[r].height = 30

chart2 = BarChart()
chart2.type = 'col'
chart2.grouping = 'clustered'
chart2.title = '各配置算法权重分布(归一化)'
chart2.y_axis.title = '权重'
data = Reference(ws, min_col=2, min_row=1, max_col=len(order) + 1, max_row=len(algo_order) + 1)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(algo_order) + 1)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.height = 9
chart2.width = 22
ws.add_chart(chart2, 'A11')

# ===== Sheet 4: 应用后验证细节 =====
ws = wb.create_sheet('应用后验证')
ws.column_dimensions['A'].width = 24
for i in range(2, 8):
    ws.column_dimensions[get_column_letter(i)].width = 13
hdr(ws, 1, ['指标', '随机基线', '静态默认A', '无约束EWMA(B)', '★G应用后', '变化(B→G)'])
applied_top1 = applied['top1_overall']
applied_top3 = applied['top3_overall']
applied_top5 = applied['top5_overall']
applied_top6 = applied['top6_overall_rate']
mc = applied['avg_top6_match_count']
b = metrics['B_ewma_unconstrained']
a = metrics['A_default']
data_rows = [
    ('Top-1%', 10.0, a['top1_overall'], b['top1_overall'], applied_top1, applied_top1 - b['top1_overall']),
    ('Top-3%', 30.0, a['top3_overall'], b['top3_overall'], applied_top3, applied_top3 - b['top3_overall']),
    ('Top-5%', 50.0, a['top5_overall'], b['top5_overall'], applied_top5, applied_top5 - b['top5_overall']),
    ('Top-6%', 60.0, a['top6_overall'], b['top6_overall'], applied_top6, applied_top6 - b['top6_overall']),
    ('match_count', 3.00, a['avg_match_count'], b['avg_match_count'], mc, round(mc - b['avg_match_count'], 3)),
]
r = 2
for name, base, va, vb, vg, d in data_rows:
    put(ws, r, 1, name, bold=True, align='left')
    put(ws, r, 2, base, color=GREEN)
    put(ws, r, 3, va, color=(RED if va < base else BLACK))
    put(ws, r, 4, vb, color=(RED if vb < base else BLACK))
    put(ws, r, 5, vg, color=(RED if vg < base else GREEN), bold=True, fillc=YELLOW)
    put(ws, r, 6, round(d, 2), color=(GREEN if d > 0 else RED), bold=True)
    r += 1
put(ws, r + 1, 1, '结论', color=BLUE, bold=True, align='left')
put(ws, r + 1, 2, 'Plan B 落地后 Top-1 较改动前无约束EWMA 回升 +1.0pp，'
                  'Top-5/6 不塌且仍超随机；自适应闭环保持安全生效。', align='left')
ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=6)
ws.row_dimensions[r + 1].height = 30

# ===== Sheet 5: 结论与下一步 =====
ws = wb.create_sheet('结论与建议')
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 90
items = [
    ('1', '根因明确：EWMA 学「覆盖命中率」，7 算法都≈随机，无区分信号 → 任何混合稀释 Top-1。'),
    ('2', 'Plan B 落地（predictor.py 三参数）：ewma_blend 0.3→0.1、minor_max_weight=0.10、ewma_alpha 可配置。'),
    ('3', '效果：应用后 Top-1 8.33%→9.33%（+1.0pp），Top-5/6 保持 52.0%/61.83%（超随机），防过均匀达成。'),
    ('4', '代价透明：自适应循环仍略逊静态默认(11.5%)；这是「学覆盖非精准」的固有局限，非bug。'),
    ('5', '建议①（Top-1优先）：ewma_blend 进一步降到 0.05，或关自适应用静态默认 — 可把 Top-1 拉回 11.5%。'),
    ('6', '建议②（根治）：把自适应评估指标从「覆盖命中率」改为「Top-1 精准度」，让 EWMA 主动 boost 频率算法。'),
    ('7', '建议③（长期）：保持验证闭环持续运行，积累 >300 期让 EWMA 更稳健，再评估是否放开 blend。'),
    ('8', '产品定位：当前为「覆盖率/缩号工具」(Top-6≈62%)，非「精准命中预测」，对外如实表述。'),
]
r = 1
put(ws, r, 1, '结论与下一步建议', color=WHITE, bold=True, fillc=BLUE)
ws.merge_cells('A1:B1')
r = 2
for n, t in items:
    put(ws, r, 1, n, bold=True, align='center', fillc=GREY)
    put(ws, r, 2, t, align='left')
    ws.row_dimensions[r].height = 32
    r += 1

os.makedirs(f'{BASE}/reports/diagnostic', exist_ok=True)
out1 = f'{BASE}/reports/diagnostic/排列5_PlanB防过均匀报告.xlsx'
out2 = f'{BASE}/reports/diagnostic/planb_report.xlsx'
wb.save(out1)
wb.save(out2)
print('已保存', out1)
print('已保存', out2)
