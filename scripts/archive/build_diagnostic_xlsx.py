# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""构建排列5 深度诊断 Excel 报告（数据诊断先行）
注：本 Windows 环境无法运行 LibreOffice 重算，故派生指标在 Python 预计算后写值，
   仍遵守配色规范：蓝=原始数据(输入) 黑=计算值 绿=基线/跨表 红=警示 黄=关键假设。
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

BASE = 'reports/diagnostic'
partA = json.load(open(f'{BASE}/partA.json', encoding='utf-8'))
iso = json.load(open(f'{BASE}/algo_isolation.json', encoding='utf-8'))
sweep = json.load(open(f'{BASE}/robustness_sweep.json', encoding='utf-8'))

POS = ['万位', '千位', '百位', '十位', '个位']
BLUE = '0000FF'; BLACK = '000000'; RED = 'FF0000'; YELLOW = 'FFFF00'; GREEN = '008000'
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=15, color='1F4E78')
SUB_FONT = Font(italic=True, size=9, color='666666')
BOLD = Font(bold=True)
CENTER = Alignment(horizontal='center', vertical='center')
WRAP = Alignment(wrap_text=True, vertical='top')
thin = Side(style='thin', color='BBBBBB')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def hdr(ws, row, headers, start_col=1):
    for j, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + j, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER; c.border = BORDER

def put(ws, r, c, v, color=None, bold=False, fill=None, fmt=None, align=None, border=True):
    cell = ws.cell(row=r, column=c, value=v)
    if color: cell.font = Font(color=color, bold=bold)
    elif bold: cell.font = BOLD
    if fill: cell.fill = PatternFill('solid', fgColor=fill)
    if fmt: cell.number_format = fmt
    if align: cell.alignment = align
    elif border: cell.alignment = CENTER
    if border: cell.border = BORDER
    return cell

# ===================== Sheet 1: 概览 =====================
ws = wb.active; ws.title = '概览'
ws['A1'] = '排列5 AI 预测系统 · 深度数据诊断报告'; ws['A1'].font = TITLE_FONT
ws['A2'] = '诊断方向：数据诊断先行（真实记录 + 单算法隔离 walk-forward + 多窗口稳健性）'; ws['A2'].font = SUB_FONT
ws['A3'] = '数据来源：p5_prediction_record(992条已验证) / p5_history_data(1009期) / 本地 walk-forward 回测'; ws['A3'].font = SUB_FONT
ws['A4'] = '检索/生成时间：2026-07-18（北京时间）'; ws['A4'].font = SUB_FONT

ws['A6'] = '⚠ 三大核心发现'; ws['A6'].font = Font(bold=True, size=12, color=RED)
findings = [
    ('1. 真实生产命中未超随机',
     '992条已验证记录(均为v3.12之前产出)标准化到Top-3后命中率27.6%~30.9%，平均命中位1.458/5(<随机1.50)。反映旧模型(含破坏性边界保护)略低于随机。'),
    ('2. v3.12"超随机"不稳健',
     '融合v3.12在80期小窗口T3=31.8%看似超随机，但在150期窗口降至T3=30.0%(=随机基线)，且5窗口扫描全在30.0~30.8%波动。优势属样本/窗口偶然。'),
    ('3. 7算法均无可靠信号',
     '单算法隔离回测：频率/遗漏/贝叶斯/趋势/马尔可夫/形态/特征全部贴近基线(T1≈10%/T3≈30%)；"噪声"算法得分不输"信号"算法，融合未增效。'),
]
r = 7
for t, d in findings:
    put(ws, r, 1, t, color=RED, bold=True, align=Alignment(vertical='top', wrap_text=True))
    put(ws, r, 2, d, align=WRAP)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 42
    r += 1
put(ws, r+1, 1, '结论：当前架构对近均匀的随机开奖基本无可榨取优势；v3.12 仅把"自伤式低于随机"修复到"等于随机"。')
ws.cell(row=r+1, column=1).font = Font(bold=True, color=BLACK)
ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=8)
ws.cell(row=r+1, column=1).alignment = WRAP
ws.row_dimensions[r+1].height = 30
ws.column_dimensions['A'].width = 22
for col in 'BCDEFGH': ws.column_dimensions[col].width = 16

# ===================== Sheet 2: 数据质量快检 =====================
ws = wb.create_sheet('数据质量快检')
ws['A1'] = '数据质量快检（992 条已验证记录）'; ws['A1'].font = TITLE_FONT
ws['A3'] = '⚠ predicted_numbers 每位置 Top-N 随版本漂移 → 历史记录不可直接横比命中率，本报告统一标准化到 Top-3'
ws['A3'].font = Font(bold=True, color=RED); ws.merge_cells('A3:E3'); ws['A3'].alignment = WRAP
ws.row_dimensions[3].height = 28
hdr(ws, 5, ['Top-N 口径', '记录数', '占比', '说明'])
topn = partA['topn_dist']; total = partA['n']
order = [0, 3, 4, 5, 6]
rr = 6
for k in order:
    if k not in topn: continue
    cnt = topn[k]
    put(ws, rr, 1, f'Top-{k}' + ('（空）' if k == 0 else ''), color=BLUE)
    put(ws, rr, 2, cnt, color=BLUE)
    put(ws, rr, 3, cnt / total, color=BLACK, fmt='0.0%')
    put(ws, rr, 4, 'v3.12 之前模型产出(旧口径)' if k != 0 else '损坏记录', align=WRAP)
    rr += 1
put(ws, rr, 1, '合计', bold=True)
put(ws, rr, 2, sum(topn[k] for k in order if k in topn), bold=True, color=BLACK)
put(ws, rr, 3, 1.0, bold=True, fmt='0.0%')
rr += 2
ws.cell(row=rr, column=1, value='其它质量指标').font = Font(bold=True, size=12, color='1F4E78'); rr += 1
hdr(ws, rr, ['指标', '数值', '解读']); rr += 1
qa = [
    ('标准化Top-3自算 vs 库存match_count 一致率', f"{partA['mc_agree']}/{total} ({partA['mc_agree']/total*100:.1f}%)", '低→版本间口径不一，库存match_count不可直接横比'),
    ('predicted_numbers 损坏记录', f"{partA['empty']}/{total} ({partA['empty']/total*100:.1f}%)", '空/非字典，需清洗或重算'),
    ('置信度 vs 命中位 相关系数 r', partA['conf_corr'], '≈0→confidence_scores 无预测价值，勿据其决策'),
]
for name, val, interp in qa:
    put(ws, rr, 1, name, color=BLUE, align=WRAP)
    put(ws, rr, 2, val, color=BLUE)
    put(ws, rr, 3, interp, align=WRAP)
    ws.row_dimensions[rr].height = 30
    rr += 1
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 16; ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 40

# ===================== Sheet 3: 真实命中率(生产) =====================
ws = wb.create_sheet('真实命中率(生产)')
ws['A1'] = '真实世界命中率（992 条已验证记录，标准化到 Top-3）'; ws['A1'].font = TITLE_FONT
ws['A2'] = '随机基线：Top-1=10% / Top-3=30% / Top-5=50%（号码近均匀分布）'; ws['A2'].font = SUB_FONT
hdr(ws, 4, ['位置', 'Top-3 真实命中率', 'Top-3 减基线(30%)', 'Top-5 命中率(对照)', 'Top-5 减基线(50%)'])
t3 = partA['pos_top3_rate']; t5 = partA['pos_top5_rate']
rr = 5
for p in POS:
    put(ws, rr, 1, p, color=BLUE)
    put(ws, rr, 2, t3[p] / 100, color=BLUE, fmt='0.0%')
    d3 = t3[p] / 100 - 0.30
    put(ws, rr, 3, d3, color=(RED if d3 < 0 else BLACK), fmt='0.0%')
    put(ws, rr, 4, t5[p] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 5, t5[p] / 100 - 0.50, color=BLACK, fmt='0.0%')
    rr += 1
put(ws, rr, 1, '平均命中位', bold=True)
put(ws, rr, 2, sum(t3.values()) / 5 / 100, bold=True, color=BLACK, fmt='0.0%')
put(ws, rr, 4, partA['avg_match_count'] / 5, bold=True, color=BLUE, fmt='0.000')
put(ws, rr, 5, partA['avg_match_count'] / 5 - 0.30, bold=True, color=BLACK, fmt='0.0%')
rr += 2
ws.cell(row=rr, column=1, value='match_count(Top-3) 分布').font = Font(bold=True, size=12, color='1F4E78'); rr += 1
hdr(ws, rr, ['命中k位', '记录数', '占比']); rr += 1
md = partA['match_dist']; dist_total = sum(md.get(str(k), 0) for k in range(6))
for k in range(6):
    cnt = md.get(str(k), 0)
    put(ws, rr, 1, k, color=BLUE)
    put(ws, rr, 2, cnt, color=BLUE)
    put(ws, rr, 3, cnt / dist_total, color=BLACK, fmt='0.0%')
    rr += 1
# 图表
chart = BarChart(); chart.type = 'col'; chart.title = '逐位置 Top-3 真实命中率 vs 随机基线(30%)'
chart.y_axis.title = '命中率'; chart.x_axis.title = '位置'
data = Reference(ws, min_col=2, min_row=4, max_row=9)
cats = Reference(ws, min_col=1, min_row=5, max_row=9)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
chart.height = 8; chart.width = 16
ws.add_chart(chart, 'G4')
ws.column_dimensions['A'].width = 14
for col in 'BCDE': ws.column_dimensions[col].width = 18

# ===================== Sheet 4: 单算法贡献 =====================
ws = wb.create_sheet('单算法贡献')
ws['A1'] = '单算法隔离 walk-forward 贡献（start=400, count=150，关闭AI）'; ws['A1'].font = TITLE_FONT
ws['A2'] = '方法：某算法权重=1、其余=0，复用 P5Predictor 真实融合逻辑；逐算法独立信号质量'; ws['A2'].font = SUB_FONT
hdr(ws, 4, ['算法', 'Top-1', 'Top-3', 'Top-5', 'score', '校准分', 'vs 融合(T3差)'])
rr = 5
fused = next(x for x in iso if x['algo'] == 'fused_v312')
for item in iso:
    is_fused = item['algo'] == 'fused_v312'
    put(ws, rr, 1, item['cn'], color=(BLUE if not is_fused else BLACK), bold=is_fused)
    put(ws, rr, 2, item['top1'] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 3, item['top3'] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 4, item['top5'] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 5, item['score'], color=BLUE, fmt='0.00')
    put(ws, rr, 6, item['calibration'], color=BLUE, fmt='0.00')
    if not is_fused:
        put(ws, rr, 7, item['top3'] / 100 - fused['top3'] / 100, color=BLACK, fmt='0.0%')
    else:
        put(ws, rr, 7, '—', color=BLACK)
    rr += 1
put(ws, rr, 1, '随机基线', color=GREEN, bold=True)
put(ws, rr, 2, 0.10, color=GREEN, fmt='0.0%')
put(ws, rr, 3, 0.30, color=GREEN, fmt='0.0%')
put(ws, rr, 4, 0.50, color=GREEN, fmt='0.0%')
for c in (5, 6, 7): put(ws, rr, c, '—', color=GREEN)
chart = BarChart(); chart.type = 'col'; chart.title = '各算法 Top-3 命中率 vs 随机基线(30%)'
data = Reference(ws, min_col=3, min_row=4, max_row=rr)
cats = Reference(ws, min_col=1, min_row=5, max_row=rr)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
chart.height = 9; chart.width = 18
ws.add_chart(chart, 'I4')
ws.column_dimensions['A'].width = 14
for col in 'BCDEFG': ws.column_dimensions[col].width = 13

# ===================== Sheet 5: 稳健性扫描 =====================
ws = wb.create_sheet('稳健性扫描')
ws['A1'] = '融合 v3.12 多窗口稳健性扫描（关闭AI，纯算法）'; ws['A1'].font = TITLE_FONT
ws['A2'] = '若各窗口 Top-3 全在 30% 附近波动，则"超随机"优势属样本偶然而非真实信号'; ws['A2'].font = SUB_FONT
hdr(ws, 4, ['窗口(起始+期数)', 'Top-1', 'Top-3', 'Top-5', 'score', '趋势'])
rr = 5; sw_start = rr
rows = [s for s in sweep if 'error' not in s]
for item in rows:
    put(ws, rr, 1, item['window'], color=BLUE)
    put(ws, rr, 2, item['top1'] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 3, item['top3'] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 4, item['top5'] / 100, color=BLUE, fmt='0.0%')
    put(ws, rr, 5, item['score'], color=BLUE, fmt='0.00')
    put(ws, rr, 6, item['trend'], color=BLUE)
    rr += 1
sw_end = rr - 1
t3s = [s['top3'] / 100 for s in rows]
put(ws, rr, 1, '跨窗口均值', bold=True, fill=YELLOW)
put(ws, rr, 2, sum([s['top1'] for s in rows]) / len(rows) / 100, bold=True, color=BLACK, fmt='0.0%')
put(ws, rr, 3, sum(t3s) / len(t3s), bold=True, color=BLACK, fmt='0.0%')
put(ws, rr, 4, sum([s['top5'] for s in rows]) / len(rows) / 100, bold=True, color=BLACK, fmt='0.0%')
put(ws, rr, 5, sum([s['score'] for s in rows]) / len(rows), bold=True, color=BLACK, fmt='0.00')
rr += 1
put(ws, rr, 1, 'Min', bold=True); put(ws, rr, 3, min(t3s), bold=True, color=BLACK, fmt='0.0%')
rr += 1
put(ws, rr, 1, 'Max', bold=True); put(ws, rr, 3, max(t3s), bold=True, color=BLACK, fmt='0.0%')
chart = LineChart(); chart.title = '融合 v3.12 各窗口 Top-3 命中率（含基线30%）'
data = Reference(ws, min_col=3, min_row=4, max_row=sw_end)
cats = Reference(ws, min_col=1, min_row=sw_start, max_row=sw_end)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
ws.cell(row=sw_end + 2, column=1, value='基线'); ws.cell(row=sw_end + 2, column=3, value=0.30)
chart.add_data(Reference(ws, min_col=3, min_row=sw_end + 2, max_row=sw_end + 2), titles_from_data=True)
chart.height = 9; chart.width = 18
ws.add_chart(chart, 'H4')
ws.column_dimensions['A'].width = 18
for col in 'BCDEF': ws.column_dimensions[col].width = 12

# ===================== Sheet 6: 结论与建议 =====================
ws = wb.create_sheet('结论与建议')
ws['A1'] = '所以呢？—— 可执行建议'; ws['A1'].font = TITLE_FONT
recs = [
    ('A. 先校准预期，停止"超随机"宣传',
     '数据清晰显示系统当前≈随机。对外/对内报告应如实标注"覆盖率工具"定位，避免基于80期小窗口的乐观结论误导决策。'),
    ('B. 用 v3.12 真跑验证闭环，补足生产证据',
     '现有992条记录全是旧模型。立即用 v3.12 生成预测并跑 verify_pending_predictions() 闭环，积累≥100条 v3.12 验证记录，确认回测优势是否落到真实世界。'),
    ('C. 修复数据质量再谈优化',
     '统一 predicted_numbers 到 Top-3 口径入库；清洗11条损坏记录；停用无校准价值的 confidence_scores 决策用途。没有干净数据，任何算法调参都是沙上筑塔。'),
    ('D. 若须继续提升，转向"新信号源"而非调权重',
     '7算法已证明无可靠边缘。再调权重收益有限。可行方向：跨期结构特征、或承认随机性将产品定位为"缩号覆盖"而非"命中预测"。'),
]
r = 3
for t, d in recs:
    put(ws, r, 1, t, color='1F4E78', bold=True, align=Alignment(vertical='top', wrap_text=True))
    put(ws, r, 2, d, align=WRAP)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    ws.row_dimensions[r].height = 48
    r += 1
ws.column_dimensions['A'].width = 24
for col in 'BCDEFGHI': ws.column_dimensions[col].width = 14

os.makedirs(BASE, exist_ok=True)
out = os.path.join(BASE, '排列5深度诊断报告.xlsx')
wb.save(out)
print('saved', out)
