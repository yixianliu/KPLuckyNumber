# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
v3.16 阶段2 (P0): 窗口参数扫参 — lookback 网格搜索

用途: 在阶段0冻结基线之上, 对 frequency_weighted.params.lookback_periods 做网格搜索,
      评估"近期窗口长度"对标准 Top-N 命中率的影响, 用 95% 置信区间判定是否稳健超越
      随机基线 / 冻结基线(方案见 reports/diagnostic/排列5_命中率优化方案.md P0节)。

复用:
  - opt_freeze_baseline.py 的评估内核(_evaluate_prediction/_calculate_overall_stats), 口径与 Backtester 一致
  - tuning_config.py 的 TuningConfig 框架(阶段1交付), 通过 to_predictor_custom_config() 注入配置
  - 读取阶段0产物 freeze_baseline_v315.json 作为对比锚点

诚实前提: 排列5 公平摇号, 历史走势无法稳定超越随机基线(Top-1≈10%/Top-5≈50%)。
          本脚本只是"系统化量化"手段, 任何"提升"结论须满足:
            ① 95%CI 下界 > 随机基线
            ② 跨 ≥3 个测试窗口(不同 test_count)稳健 —— 本脚本单窗口, 仅给出候选, 不判最终胜出

用法:
    python opt_tune_sweep.py                 # 默认 lookback 网格, 测试最近 300 期
    python opt_tune_sweep.py 200             # 测试最近 200 期
    python opt_tune_sweep.py 300 30,40,60,90 # 自定义网格
"""
import sys
import os
import json
import copy
import logging
from datetime import datetime

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logging.getLogger().setLevel(logging.WARNING)

# 复用阶段0评估内核(无 matplotlib 依赖, 安全 import)
from opt_freeze_baseline import (
    _evaluate_prediction, _calculate_overall_stats, _config_fingerprint,
    POSITION_NAMES, RANDOM_BASELINE,
)
from modules.predictor import P5Predictor, P5PredictorConfig
from modules.database import P5Database
from tuning_config import _control_template, TuningConfig

DEFAULT_GRID = [30, 40, 50, 60, 70, 80, 90]
FREEZE_BASELINE_PATH = 'reports/backtest/freeze_baseline_v315.json'


# ----------------------------------------------------------------------------
# walk-forward 内核(与阶段0同口径, 接受自定义 predictor)
# ----------------------------------------------------------------------------
def run_walk_forward(predictor, test_count=300, start_index_min=50):
    db = P5Database()
    db.connect()
    history = db.get_history_data(limit=None, order='ASC')
    total_hist = len(history)
    db.disconnect()

    start_index = max(start_index_min, total_hist - test_count)
    effective = min(test_count, total_hist - start_index)

    results = []
    for i in range(start_index, start_index + effective):
        train = history[:i]
        target = history[i]['issue']
        actual = history[i]['numbers']
        pr = predictor.predict(train, target)
        if 'error' in pr:
            continue
        results.append(_evaluate_prediction(pr, actual, target))

    agg, ci = _calculate_overall_stats(results)
    return {
        'meta': {
            'test_count_effective': effective,
            'start_index': start_index,
            'total_history': total_hist,
        },
        'overall_stats': agg,
        'confidence_95': ci,
        'fingerprint': _config_fingerprint(predictor),
    }


def _build_predictor_for_lookback(lookback):
    """用阶段1框架生成控制组配置, 仅改 lookback, 注入预测器(禁用AI)"""
    tpl = _control_template()
    tpl['algorithms']['frequency_weighted']['params']['lookback_periods'] = lookback
    tc = TuningConfig(tpl)
    ok, errs = tc.validate()
    if not ok:
        raise ValueError(f'lookback={lookback} 配置校验失败: {errs}')
    custom = tc.to_predictor_custom_config()
    cfg = P5PredictorConfig(custom)
    predictor = P5Predictor(cfg)
    predictor.ai_available = False
    predictor.config.config.setdefault('global', {})['enable_ai_model'] = False
    return predictor


# ----------------------------------------------------------------------------
# 对比与判定
# ----------------------------------------------------------------------------
def compare_to_baseline(sweep_result, baseline):
    ci = sweep_result['confidence_95']
    base_ci = baseline['confidence_95']
    base_agg = baseline['overall_stats']
    out = {}
    for k in ('top1', 'top3', 'top5', 'top6'):
        rate = ci[k]['rate']
        random = RANDOM_BASELINE[k]
        base_rate = base_agg[f'avg_{k}_hit_rate']
        out[k] = {
            'rate': rate,
            'ci95_low': ci[k]['ci95_low'],
            'ci95_high': ci[k]['ci95_high'],
            'random': random,
            'baseline_rate': base_rate,
            'baseline_ci95_low': base_ci[k]['ci95_low'],
            'deviation_vs_random': round(rate - random, 2),
            'deviation_vs_baseline': round(rate - base_rate, 2),
            'beats_random': ci[k]['ci95_low'] > random,
            'beats_baseline': ci[k]['ci95_low'] > base_ci[k]['ci95_low'],
        }
    return out


# ----------------------------------------------------------------------------
# Excel 仪表盘
# ----------------------------------------------------------------------------
def build_excel(grid, sweep_results, baseline, comparisons, path):
    wb = Workbook()
    C_BLUE = '1F4E78'
    C_HEAD = '2E5496'
    C_YELLOW = 'FFF2CC'
    C_RED = 'C00000'
    C_GREEN = '375623'
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ---- Sheet1 扫参总览 ----
    ws = wb.active
    ws.title = '扫参总览'
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'v3.16 P0 窗口扫参 — lookback 对标准 Top-N 命中率的影响'
    ws['A1'].font = Font(size=14, bold=True, color=C_BLUE)
    ws.merge_cells('A1:I1')
    ws['A2'] = (f"生成时间: {datetime.now().isoformat()}  |  测试期数: {sweep_results[grid[0]]['meta']['test_count_effective']}  |  "
                f"对比锚点: freeze_baseline_v315.json")
    ws['A2'].font = Font(size=9, italic=True, color='808080')
    ws.merge_cells('A2:I2')

    headers = ['lookback', 'Top-1(%)', 'Top-3(%)', 'Top-5(%)', 'Top-6(%)',
               'Top-5 CI下限', 'Top-5 超越随机?', 'Top-5 超越基线?', 'vs基线偏差']
    r0 = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border

    for i, L in enumerate(grid):
        row = r0 + 1 + i
        comp = comparisons[L]
        vals = [
            L, comp['top1']['rate'], comp['top3']['rate'], comp['top5']['rate'], comp['top6']['rate'],
            comp['top5']['ci95_low'],
            '是' if comp['top5']['beats_random'] else '否',
            '是' if comp['top5']['beats_baseline'] else '否',
            comp['top5']['deviation_vs_baseline'],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 1:
                cell.font = Font(bold=True)
            if c == 7:
                cell.font = Font(bold=True, color=C_GREEN if comp['top5']['beats_random'] else C_RED)
            if c == 8:
                cell.font = Font(bold=True, color=C_GREEN if comp['top5']['beats_baseline'] else C_RED)
            if c == 6:
                cell.fill = PatternFill('solid', fgColor=C_YELLOW)

    # 数据条(对 Top-5 命中率列)
    last_row = r0 + len(grid)
    ws.conditional_formatting.add(
        f'D{r0+1}:D{last_row}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color='5B9BD5', showValue=True))

    # 随机基线标注行
    base_row = last_row + 1
    ws.cell(row=base_row, column=1, value='随机基线').font = Font(bold=True, color=C_RED)
    for c, k in zip(range(2, 6), ('top1', 'top3', 'top5', 'top6')):
        cell = ws.cell(row=base_row, column=c, value=RANDOM_BASELINE[k])
        cell.fill = PatternFill('solid', fgColor=C_YELLOW)
        cell.border = border
        cell.alignment = center
    # 冻结基线标注行
    fb_row = base_row + 1
    ws.cell(row=fb_row, column=1, value='冻结基线(v3.15)').font = Font(bold=True, color=C_BLUE)
    base_agg = baseline['overall_stats']
    for c, k in zip(range(2, 6), ('top1', 'top3', 'top5', 'top6')):
        cell = ws.cell(row=fb_row, column=c, value=base_agg[f'avg_{k}_hit_rate'])
        cell.border = border
        cell.alignment = center

    # 结论
    concl_row = fb_row + 2
    any_beat = any(comparisons[L]['top5']['beats_random'] for L in grid)
    best = max(grid, key=lambda L: comparisons[L]['top5']['rate'])
    conclusion = (
        f"结论: 7 档 lookback 中, Top-5 最高为 lookback={best} ({comparisons[best]['top5']['rate']:.2f}%), "
        f"但所有档位的 95%CI 下限均 ≤ 随机基线(50%)"
        if not any_beat else
        f"结论: 存在 lookback 档位 Top-5 的 95%CI 下限 > 随机基线, 但单窗口结果需跨≥3测试窗口复核确认稳健性。"
    )
    ws.cell(row=concl_row, column=1, value=conclusion)
    ws.cell(row=concl_row, column=1).font = Font(bold=True, color=C_BLUE)
    ws.cell(row=concl_row, column=1).alignment = left
    ws.merge_cells(start_row=concl_row, start_column=1, end_row=concl_row, end_column=9)
    ws.row_dimensions[concl_row].height = 45

    widths = [12, 11, 11, 11, 11, 14, 16, 16, 14]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet2 全口径对比(逐 lookback × 4 指标 vs 基线/随机) ----
    ws2 = wb.create_sheet('全口径对比')
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = '各 lookback 档位 × Top-N 命中率 + 偏差(随机/基线)'
    ws2['A1'].font = Font(size=13, bold=True, color=C_BLUE)
    ws2.merge_cells('A1:L1')

    h2 = ['lookback', 'Top-1', 'vs随机', 'vs基线', 'Top-3', 'vs随机', 'vs基线',
          'Top-5', 'vs随机', 'vs基线', 'Top-6', 'vs随机']
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    for i, L in enumerate(grid):
        row = 4 + i
        comp = comparisons[L]
        vals = [
            L,
            comp['top1']['rate'], comp['top1']['deviation_vs_random'], comp['top1']['deviation_vs_baseline'],
            comp['top3']['rate'], comp['top3']['deviation_vs_random'], comp['top3']['deviation_vs_baseline'],
            comp['top5']['rate'], comp['top5']['deviation_vs_random'], comp['top5']['deviation_vs_baseline'],
            comp['top6']['rate'], comp['top6']['deviation_vs_random'],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 1:
                cell.font = Font(bold=True)
    for c, w in enumerate([12, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9, 9], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # 折线图: Top-5 随 lookback 变化
    chart = LineChart()
    chart.title = 'Top-5 命中率 vs lookback'
    chart.y_axis.title = '命中率(%)'
    chart.x_axis.title = 'lookback'
    data = Reference(ws2, min_col=8, min_row=3, max_row=3 + len(grid))  # Top-5 列
    cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(grid))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws2.add_chart(chart, 'N3')

    wb.save(path)


# ----------------------------------------------------------------------------
# 主流程
# ----------------------------------------------------------------------------
def main():
    test_count = int(sys.argv[1]) if len(sys.argv) > 1 else 300
    grid = DEFAULT_GRID
    if len(sys.argv) > 2:
        grid = [int(x) for x in sys.argv[2].split(',') if x.strip()]
    grid = sorted(set(grid))

    # 载入阶段0冻结基线
    if not os.path.exists(FREEZE_BASELINE_PATH):
        raise SystemExit(f'缺少阶段0产物 {FREEZE_BASELINE_PATH}，请先运行 opt_freeze_baseline.py')
    with open(FREEZE_BASELINE_PATH, encoding='utf-8') as f:
        baseline = json.load(f)

    print(f'>>> P0 窗口扫参: lookback={grid}, 测试最近 {test_count} 期, 对比冻结基线')
    sweep_results = {}
    comparisons = {}
    for L in grid:
        print(f'  · lookback={L} ...', end=' ', flush=True)
        predictor = _build_predictor_for_lookback(L)
        res = run_walk_forward(predictor, test_count=test_count)
        sweep_results[L] = res
        comparisons[L] = compare_to_baseline(res, baseline)
        c = comparisons[L]['top5']
        print(f"Top-5={c['rate']:.2f}% (CI[{c['ci95_low']},{c['ci95_high']}]) "
              f"超越随机={'是' if c['beats_random'] else '否'} 超越基线={'是' if c['beats_baseline'] else '否'}")

    os.makedirs('reports/backtest', exist_ok=True)
    out = {
        'meta': {
            'generated_at': datetime.now().isoformat(),
            'script': 'opt_tune_sweep.py',
            'purpose': 'v3.16 P0: frequency_weighted lookback 窗口扫参',
            'grid': grid,
            'test_count': test_count,
            'baseline_path': FREEZE_BASELINE_PATH,
            'honest_note': '排列5公平摇号, 单窗口扫参仅给候选; 最终胜出须跨≥3测试窗口稳健且CI下界>随机',
        },
        'random_baseline': RANDOM_BASELINE,
        'baseline_overall_stats': baseline['overall_stats'],
        'sweep': {
            str(L): {
                'meta': sweep_results[L]['meta'],
                'overall_stats': sweep_results[L]['overall_stats'],
                'confidence_95': sweep_results[L]['confidence_95'],
                'fingerprint': sweep_results[L]['fingerprint'],
                'comparison': comparisons[L],
            } for L in grid
        },
    }
    json_path = 'reports/backtest/sweep_lookback_p0.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f'>>> JSON 已保存: {json_path}')

    xlsx_path = 'reports/backtest/sweep_lookback_p0.xlsx'
    build_excel(grid, sweep_results, baseline, comparisons, xlsx_path)
    print(f'>>> Excel 已保存: {xlsx_path}')

    # 汇总
    best = max(grid, key=lambda L: comparisons[L]['top5']['rate'])
    print(f'\n=== P0 扫参汇总 (测试{test_count}期) ===')
    print(f'  最优 lookback(按Top-5): {best} → {comparisons[best]["top5"]["rate"]:.2f}%')
    any_beat = any(comparisons[L]['top5']['beats_random'] for L in grid)
    print(f'  是否有档位 Top-5 的 95%CI 下界 > 随机基线(50%): {"是" if any_beat else "否(全部在基线内)"}')


if __name__ == '__main__':
    main()
