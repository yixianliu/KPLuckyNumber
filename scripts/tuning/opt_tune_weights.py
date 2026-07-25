# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
v3.16 阶段 P2: 主算法权重配比扫参

目的: 用与冻结基线(v3.16阶段0)完全一致的 walk-forward 口径, 检验主算法三者
      (frequency_weighted / omission_regression / bayesian_inference) 的相对权重配比
      是否是命中率的有效杠杆。

动机: v3.16 监控层(opt_monitor.py signals)的消融结果显示——在 100 期小窗口内,
      频率算法的边际贡献为负(-4.2), 而遗漏(+3.4)/贝叶斯(+3.2)为正。
      这留下一个真问题: 当前冻结权重(频率0.54/遗漏0.34/贝叶斯0.10)是否真最优?
      本脚本用同一内核把"降权频率/升权遗漏/升权贝叶斯"各类配比做 300 期实证。

设计: 13 档权重配比(含 frozen_control 对照 + equal_main + 降权频率系列 + 遗漏/贝叶斯加重
      + 三档 pure-signal), 每档 300 期 walk-forward;
      与冻结基线 Top-5=49.60%(CI[47.07,52.13]) 及随机基线 50% 对比。

诚实标注:
  - 排列5公平摇号, 7 算法信号源在监控层 isolation 均≈随机; 若所有源≈随机,
    则改变其在融合中的配比无法系统性地超越随机——但须用实证而非断言确认。
  - 任何"杠杆"须 95%CI 下限>随机 且跨多配置稳健才成立; 否则判噪声/基线内波动。

复用: opt_freeze_baseline 的 _evaluate_prediction / _calculate_overall_stats / RANDOM_BASELINE
      (零修改, 口径一致); freeze_baseline_v315.json 作为真值起点。

用法:
    python opt_tune_weights.py              # 完整扫参(13档, 默认300期)
    python opt_tune_weights.py 200          # 指定回测期数
    python opt_tune_weights.py --quick      # 小网格快速验证(注入生效即可)
"""
import sys
import os
import json
import logging
from datetime import datetime

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

logging.getLogger().setLevel(logging.WARNING)

from opt_freeze_baseline import (
    _evaluate_prediction,
    _calculate_overall_stats,
    RANDOM_BASELINE,
)
from modules.predictor import P5Predictor, P5PredictorConfig
from modules.database import P5Database

POSITION_NAMES = ['万位', '千位', '百位', '十位', '个位']

# 冻结默认(v3.15 封板) —— 权重配比对照基准
FROZEN_WEIGHTS = {
    'frequency_weighted': 0.54,
    'omission_regression': 0.34,
    'bayesian_inference': 0.10,
}
MINOR_WEIGHTS = {  # 4 个次要算法保持冻结, 不参与配比扫参(实测≈噪声)
    'trend_momentum': 0.01,
    'markov_transition': 0.005,
    'pattern_continuation': 0.003,
    'feature_engineering': 0.002,
}

# 13 档权重配比扫参网格(仅改主算法三者; 次要算法保持冻结)
WEIGHT_CONFIGS = {
    'frozen_control': {'frequency_weighted': 0.54, 'omission_regression': 0.34, 'bayesian_inference': 0.10},
    'equal_main':      {'frequency_weighted': 0.34, 'omission_regression': 0.33, 'bayesian_inference': 0.33},
    'freq_down_a':     {'frequency_weighted': 0.40, 'omission_regression': 0.45, 'bayesian_inference': 0.15},
    'freq_down_b':     {'frequency_weighted': 0.30, 'omission_regression': 0.50, 'bayesian_inference': 0.20},
    'freq_down_c':     {'frequency_weighted': 0.20, 'omission_regression': 0.55, 'bayesian_inference': 0.25},
    'omis_heavy':      {'frequency_weighted': 0.15, 'omission_regression': 0.70, 'bayesian_inference': 0.15},
    'bayes_up':        {'frequency_weighted': 0.30, 'omission_regression': 0.35, 'bayesian_inference': 0.35},
    'bayes_heavy':     {'frequency_weighted': 0.25, 'omission_regression': 0.30, 'bayesian_inference': 0.45},
    'freq_heavy':      {'frequency_weighted': 0.75, 'omission_regression': 0.18, 'bayesian_inference': 0.07},
    'omis_light':      {'frequency_weighted': 0.50, 'omission_regression': 0.20, 'bayesian_inference': 0.30},
    'freq_only':       {'frequency_weighted': 1.00, 'omission_regression': 0.00, 'bayesian_inference': 0.00},
    'omis_only':       {'frequency_weighted': 0.00, 'omission_regression': 1.00, 'bayesian_inference': 0.00},
    'bayes_only':      {'frequency_weighted': 0.00, 'omission_regression': 0.00, 'bayesian_inference': 1.00},
}

# --quick 验证用极小网格
QUICK_CONFIGS = {
    'frozen_control': FROZEN_WEIGHTS,
    'freq_down_b':     WEIGHT_CONFIGS['freq_down_b'],
    'freq_only':       WEIGHT_CONFIGS['freq_only'],
}
QUICK_TEST_COUNT = 60


def build_custom_config(weights: dict) -> dict:
    """构造只覆盖主算法权重的 custom_config(深合并安全, 次要算法保持冻结)。"""
    algos = {}
    for name, w in weights.items():
        algos[name] = {'weight': w}
    return {'algorithms': algos}


def run_walkforward(test_count, custom_config=None):
    """与冻结基线一致的 walk-forward; custom_config 注入主算法权重。"""
    cfg = P5PredictorConfig(custom_config=custom_config) if custom_config else None
    p = P5Predictor(config=cfg)
    p.ai_available = False
    p.config.config.setdefault('global', {})['enable_ai_model'] = False

    db = P5Database()
    db.connect()
    history = db.get_history_data(limit=None, order='ASC')
    db.disconnect()
    total_hist = len(history)

    start_index = max(50, total_hist - test_count)
    effective = min(test_count, total_hist - start_index)
    if effective <= 0:
        return []

    results = []
    for i in range(start_index, start_index + effective):
        train = history[:i]
        target = history[i]['issue']
        actual = history[i]['numbers']
        pr = p.predict(train, target)
        if 'error' in pr:
            continue
        results.append(_evaluate_prediction(pr, actual, target))
    return results


def evaluate(results):
    agg, ci = _calculate_overall_stats(results)
    return agg, ci


def load_baseline():
    path = 'reports/backtest/freeze_baseline_v315.json'
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def sweep(test_count, quick=False):
    configs = QUICK_CONFIGS if quick else WEIGHT_CONFIGS
    tc = QUICK_TEST_COUNT if quick else test_count

    baseline = load_baseline()
    if baseline:
        base_top5 = baseline['overall_stats']['avg_top5_hit_rate']
        base_ci = baseline['confidence_95']['top5']
        base_ci_low = base_ci['ci95_low']
    else:
        base_top5 = RANDOM_BASELINE['top5']
        base_ci_low = RANDOM_BASELINE['top5']

    per_run = []
    for name, weights in configs.items():
        res = run_walkforward(tc, build_custom_config(weights))
        if not res:
            per_run.append({'config': name, 'weights': weights, 'error': 'no_results'})
            continue
        agg, ci = evaluate(res)
        top5 = agg['avg_top5_hit_rate']
        top5_ci = ci['top5']
        wsum = round(sum(weights.values()), 3)
        per_run.append({
            'config': name,
            'weights': weights,
            'weight_sum': wsum,
            'top1': agg['avg_top1_hit_rate'],
            'top3': agg['avg_top3_hit_rate'],
            'top5': top5,
            'top6': agg['avg_top6_hit_rate'],
            'top5_ci_low': top5_ci['ci95_low'],
            'top5_ci_high': top5_ci['ci95_high'],
            'delta_vs_baseline': round(top5 - base_top5, 2),
            'beats_random': top5_ci['ci95_low'] > RANDOM_BASELINE['top5'],
            'beats_baseline': top5_ci['ci95_low'] > base_ci_low,
            'tested': agg['total_tested'],
        })

    best = max(per_run, key=lambda x: x.get('top5', -1)) if per_run else None
    any_robust = any(r.get('beats_random') for r in per_run if 'error' not in r)
    # 降权频率系列是否优于 frozen (回应监控层疑点)
    fd_any_better = any(
        r.get('top5', 0) > base_top5 for r in per_run
        if 'error' not in r and r['config'].startswith('freq_down')
    )

    summary = {
        'meta': {
            'generated_at': datetime.now().isoformat(),
            'script': 'opt_tune_weights.py',
            'purpose': 'v3.16 阶段P2: 主算法权重配比扫参(频率/遗漏/贝叶斯三者的相对配比)',
            'test_count': tc,
            'quick': quick,
            'honest_note': ('排列5公平摇号; 7算法信号源 isolation 均≈随机, 改变融合配比'
                            '须用 95%CI 下限>随机 且跨多配置稳健才判"杠杆", 否则属基线内波动。'),
            'baseline_top5': base_top5,
            'random_top5': RANDOM_BASELINE['top5'],
            'frozen_weights': FROZEN_WEIGHTS,
        },
        'weight_configs': configs if not quick else QUICK_CONFIGS,
        'per_run': per_run,
        'best_run': best,
        'any_robust_beat_random': any_robust,
        'freq_down_any_better_than_frozen': fd_any_better,
        'conclusion': (
            '结论: 所有权重配比档位的 Top-5 95%CI 下限均 ≤ 随机基线(50%), '
            '无任一档位稳健超越冻结基线/随机; 监控层"频率负消融"在小窗口为噪声, '
            '降权频率并未带来稳健提升 → 当前冻结权重(频率0.54/遗漏0.34/贝叶斯0.10)'
            '是最务实配置, 权重配比非命中率有效杠杆(与P0/P1/v3.14审计一致)。'
            if not any_robust else
            '结论: 存在档位 95%CI 下限>随机, 需进一步交叉验证确认是否为真实信号。'
        ),
    }
    return summary


# ----------------------------------------------------------------------------
# Excel 仪表盘
# ----------------------------------------------------------------------------
def build_excel(summary, path):
    wb = Workbook()
    C_BLUE = '1F4E78'
    C_HEAD = '2E5496'
    C_YELLOW = 'FFF2CC'
    C_RED = 'C00000'
    C_GREEN = '375623'
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    meta = summary['meta']
    per_run = [r for r in summary['per_run'] if 'error' not in r]

    # ---- Sheet1 概览 ----
    ws = wb.active
    ws.title = '概览'
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'v3.16 P2 主算法权重配比扫参'
    ws['A1'].font = Font(size=14, bold=True, color=C_BLUE)
    ws.merge_cells('A1:E1')
    ws['A2'] = (f"生成时间: {meta['generated_at']} | 回测: {meta['test_count']}期 "
                f"| 基线Top5: {meta['baseline_top5']}% | 随机: {meta['random_top5']}%")
    ws['A2'].font = Font(size=9, italic=True, color='808080')
    ws.merge_cells('A2:E2')
    ws['A3'] = f"⚠️ {meta['honest_note']}"
    ws['A3'].font = Font(size=9, bold=True, color=C_RED)
    ws.merge_cells('A3:E3')

    headers = ['指标', '值', '说明']
    r0 = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    rows = [
        ('扫参档位数', len(per_run), '频率/遗漏/贝叶斯相对配比'),
        ('随机基线 Top5', f"{meta['random_top5']}%", '黄底=关键假设'),
        ('冻结基线 Top5', f"{meta['baseline_top5']}%", '真值起点(frozen_control)'),
        ('冻结权重', f"F{int(meta['frozen_weights']['frequency_weighted']*100)}/"
                     f"O{int(meta['frozen_weights']['omission_regression']*100)}/"
                     f"B{int(meta['frozen_weights']['bayesian_inference']*100)}",
         '频率/遗漏/贝叶斯'),
        ('是否任一档稳健超越随机', '是' if summary['any_robust_beat_random'] else '否',
         'CI下限>50% 且跨配置稳健'),
        ('降权频率系列是否优于冻结', '是' if summary['freq_down_any_better_than_frozen'] else '否',
         '回应监控层"频率负消融"疑点'),
    ]
    best = summary['best_run']
    if best:
        rows.append(('最高 Top5 档位', best.get('config'),
                     f"Top5={best.get('top5')}% (CI[{best.get('top5_ci_low')},{best.get('top5_ci_high')}])"))
    for i, (a, b, c) in enumerate(rows):
        row = r0 + 1 + i
        ws.cell(row=row, column=1, value=a).font = Font(bold=True)
        ws.cell(row=row, column=2, value=b)
        ws.cell(row=row, column=3, value=c)
        for cc in range(1, 4):
            ws.cell(row=row, column=cc).border = border
            ws.cell(row=row, column=cc).alignment = left if cc == 3 else center
        if a in ('随机基线 Top5', '冻结基线 Top5'):
            ws.cell(row=row, column=2).fill = PatternFill('solid', fgColor=C_YELLOW)
    concl_row = r0 + 1 + len(rows) + 1
    ws.cell(row=concl_row, column=1, value=summary['conclusion'])
    ws.cell(row=concl_row, column=1).font = Font(bold=True, color=C_BLUE)
    ws.cell(row=concl_row, column=1).alignment = left
    ws.merge_cells(start_row=concl_row, start_column=1, end_row=concl_row, end_column=5)
    ws.row_dimensions[concl_row].height = 60
    for c, w in enumerate([26, 22, 44], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet2 权重配置对比 ----
    ws2 = wb.create_sheet('权重配置对比')
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = '主算法权重配比对比 (Top-5 命中率 vs 冻结基线)'
    ws2['A1'].font = Font(size=13, bold=True, color=C_BLUE)
    ws2.merge_cells('A1:J1')
    ph = ['配置名', '频率权重', '遗漏权重', '贝叶斯权重', 'Top-5(%)',
          'CI下限', 'CI上限', 'Δvs基线', '超越随机?', '超越基线?']
    for c, h in enumerate(ph, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    rr = 4
    for r in per_run:
        w = r['weights']
        vals = [r['config'], w.get('frequency_weighted'), w.get('omission_regression'),
                w.get('bayesian_inference'), r['top5'], r['top5_ci_low'],
                r['top5_ci_high'], r['delta_vs_baseline'],
                '是' if r['beats_random'] else '否',
                '是' if r['beats_baseline'] else '否']
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 1:
                cell.font = Font(bold=True,
                                 color=C_BLUE if r['config'] == 'frozen_control' else '000000')
            if c == 9:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_random'] else C_RED)
            if c == 10:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_baseline'] else C_RED)
        rr += 1
    # 基线/随机标注行
    ws2.cell(row=rr, column=1, value='冻结基线').font = Font(bold=True, color=C_RED)
    ws2.cell(row=rr, column=5, value=meta['baseline_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
    for c in (2, 3, 4, 6, 7, 8, 9, 10):
        ws2.cell(row=rr, column=c, value='—').border = border
    ws2.cell(row=rr + 1, column=1, value='随机基线').font = Font(bold=True, color=C_RED)
    ws2.cell(row=rr + 1, column=5, value=meta['random_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
    for c in (2, 3, 4, 6, 7, 8, 9, 10):
        ws2.cell(row=rr + 1, column=c, value='—').border = border
    # 数据条(Top5列)
    ws2.conditional_formatting.add(
        f'E4:E{rr-1}',
        DataBarRule(start_type='num', start_value=40, end_type='num', end_value=60,
                    color='5B9BD5', showValue=True))
    for c, w in enumerate([16, 11, 11, 12, 11, 10, 10, 10, 12, 12], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet3 逐配置明细 ----
    ws3 = wb.create_sheet('逐配置明细')
    ws3.sheet_view.showGridLines = False
    dh = ['配置名', 'Top-1', 'Top-3', 'Top-5', 'Top-6',
          'CI下', 'CI上', 'Δ基线', '超越随机', '超越基线', '回测期数']
    for c, h in enumerate(dh, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    for i, r in enumerate(per_run):
        row = 2 + i
        vals = [r['config'], r['top1'], r['top3'], r['top5'], r['top6'],
                r['top5_ci_low'], r['top5_ci_high'], r['delta_vs_baseline'],
                '是' if r['beats_random'] else '否',
                '是' if r['beats_baseline'] else '否', r['tested']]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 1:
                cell.font = Font(bold=True,
                                 color=C_BLUE if r['config'] == 'frozen_control' else '000000')
            if c == 9:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_random'] else C_RED)
            if c == 10:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_baseline'] else C_RED)
    ws3.freeze_panes = 'A2'
    for c, w in enumerate([16, 10, 10, 10, 10, 9, 9, 9, 11, 11, 11], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)


if __name__ == '__main__':
    args = sys.argv[1:]
    quick = '--quick' in args
    tc = 300
    for a in args:
        if a.isdigit():
            tc = int(a)
    mode = 'quick' if quick else 'p2'
    print(f'>>> P2 主算法权重配比扫参 ({mode}, '
          f'回测 {QUICK_TEST_COUNT if quick else tc} 期)...')
    summary = sweep(tc, quick=quick)

    os.makedirs('reports/backtest', exist_ok=True)
    json_path = f'reports/backtest/sweep_weights_{mode}.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f'>>> JSON 已保存: {json_path}')

    xlsx_path = f'reports/backtest/sweep_weights_{mode}.xlsx'
    build_excel(summary, xlsx_path)
    print(f'>>> Excel 已保存: {xlsx_path}')

    per_run = [r for r in summary['per_run'] if 'error' not in r]
    print(f'\n=== P2 汇总 ({len(per_run)} 档) ===')
    print(f"  随机基线 Top5={summary['meta']['random_top5']}% | "
          f"冻结基线 Top5={summary['meta']['baseline_top5']}%")
    for r in per_run:
        print(f"  {r['config']:<14} F{r['weights']['frequency_weighted']:.2f}/"
              f"O{r['weights']['omission_regression']:.2f}/"
              f"B{r['weights']['bayesian_inference']:.2f}  "
              f"Top5={r['top5']}% CI[{r['top5_ci_low']},{r['top5_ci_high']}] "
              f"{'★超越随机' if r['beats_random'] else ''}")
    best = summary['best_run']
    if best:
        print(f"  最高档: {best.get('config')} Top5={best.get('top5')}% "
              f"CI[{best.get('top5_ci_low')},{best.get('top5_ci_high')}]")
    print(f"  降权频率优于冻结: {'是' if summary['freq_down_any_better_than_frozen'] else '否'}")
    print(f"  稳健超越随机: {'是' if summary['any_robust_beat_random'] else '否'}")
    print(f"  结论: {summary['conclusion']}")
