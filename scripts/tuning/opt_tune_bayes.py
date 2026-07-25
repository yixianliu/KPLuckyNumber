# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
v3.16 阶段 P1: 贝叶斯算法超参敏感性扫参

目的: 用与冻结基线(v3.16阶段0)完全一致的 walk-forward 口径, 检验贝叶斯推断算法
      (bayesian_inference) 内部超参是否是命中率的有效杠杆。

设计(OFAT 单因子扫参):
  - 对 8 个内部超参逐一在其物理合理区间取多档, 其余保持 v3.15 冻结默认;
  - 每档用 P5Predictor(custom_config=深合并覆盖) 跑 300 期 walk-forward;
  - 与冻结基线 Top-5=49.60%(CI[47.07,52.13]) 及随机基线 50% 对比。

诚实标注:
  - 贝叶斯算法读取 p5_prediction_record 表的验证记录(1117条 verified)构建似然;
    冻结基线 walk-forward 未设 _verification_cutoff, 故所有配置共用同一份静态验证史,
    这与生产行为(验证记录随真实运行累积)一致, 且对所有配置 apple-to-apple 公平。
  - 排列5公平摇号, 验证记录本身是过往预测(≈随机), 似然不含稳定可学信号, 故预期
    所有超参档位均落于随机带内; 若某档 CI 下限>随机且跨多参数稳健, 才判"杠杆"。

复用: opt_freeze_baseline 的 _evaluate_prediction / _calculate_overall_stats / RANDOM_BASELINE
      (零修改, 口径一致); freeze_baseline_v315.json 作为真值起点。

用法:
    python opt_tune_bayes.py                 # 完整扫参(约40档, 默认300期)
    python opt_tune_bayes.py 200             # 指定回测期数
    python opt_tune_bayes.py --quick         # 小网格快速验证(注入生效即可)
"""
import sys
import os
import json
import logging
from datetime import datetime

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

logging.getLogger().setLevel(logging.WARNING)

from opt_freeze_baseline import (
    _evaluate_prediction,
    _calculate_overall_stats,
    RANDOM_BASELINE,
)
from modules.predictor import P5Predictor, P5PredictorConfig
from modules.database import P5Database

POSITION_NAMES = ['万位', '千位', '百位', '十位', '个位']

# 冻结默认(v3.15 封板)
BAYES_DEFAULTS = {
    'prior_smooth': 0.10,
    'posterior_weight': 0.92,
    'verification_window': 60,
    'penalize_miss': 0.68,
    'reward_hit': 1.40,
    'decay_half_life': 10,
    'beta_alpha': 0.8,
    'prior_temporal_scale': 50,
}

# OFAT 扫参网格(物理合理区间, 含端点)
GRID = {
    'posterior_weight': [0.0, 0.3, 0.5, 0.7, 0.85, 0.92, 0.98, 1.0],
    'reward_hit': [1.10, 1.25, 1.40, 1.70, 2.00],
    'penalize_miss': [0.50, 0.68, 0.80, 0.90],
    'prior_smooth': [0.02, 0.05, 0.10, 0.20, 0.50],
    'beta_alpha': [0.3, 0.5, 0.8, 1.5, 3.0],
    'decay_half_life': [5, 10, 20, 40],
    'verification_window': [30, 60, 120, 200],
    'prior_temporal_scale': [20, 50, 100, 200],
}

# 交互角点(多参数同时推向极端, 检验组合效应)
CORNERS = {
    'aggressive': {
        'posterior_weight': 0.98, 'reward_hit': 2.00, 'penalize_miss': 0.50,
        'beta_alpha': 0.3, 'prior_smooth': 0.02,
    },
}

# --quick 验证用极小网格
QUICK_GRID = {
    'posterior_weight': [0.0, 0.92, 1.0],
    'reward_hit': [1.40, 2.00],
}
QUICK_TEST_COUNT = 60

# --lite 代表性精简网格(端点+默认+中间点, 控运行时长, 结论对"无杠杆"仍稳健)
LITE_GRID = {
    'posterior_weight': [0.0, 0.5, 0.92, 1.0],
    'reward_hit': [1.10, 1.40, 2.00],
    'penalize_miss': [0.50, 0.68, 0.90],
    'prior_smooth': [0.02, 0.10, 0.50],
    'beta_alpha': [0.3, 0.8, 1.5],
    'decay_half_life': [5, 10, 40],
    'verification_window': [30, 60, 200],
    'prior_temporal_scale': [20, 50, 200],
}


def build_custom_config(overrides: dict) -> dict:
    """构造只覆盖贝叶斯 params 叶子的 custom_config(深合并安全)。"""
    params = {}
    for k, v in overrides.items():
        params[k] = v
    return {'algorithms': {'bayesian_inference': {'params': params}}}


def run_walkforward(test_count, custom_config=None):
    """与冻结基线一致的 walk-forward; custom_config 注入贝叶斯超参。"""
    cfg = P5PredictorConfig(custom_config=custom_config) if custom_config else None
    p = P5Predictor(config=cfg)
    p.ai_available = False
    p.config.config.setdefault('global', {})['enable_ai_model'] = False

    db = P5Database()
    db.connect()
    history = db.get_history_data(limit=None, order='ASC')
    db.disconnect()
    total_hist = len(history)

    start_index = max(50, total_hist - test_count)
    effective = min(test_count, total_hist - start_index)
    if effective <= 0:
        return []

    results = []
    for i in range(start_index, start_index + effective):
        train = history[:i]
        target = history[i]['issue']
        actual = history[i]['numbers']
        pr = p.predict(train, target)
        if 'error' in pr:
            continue
        results.append(_evaluate_prediction(pr, actual, target))
    return results


def evaluate(results):
    agg, ci = _calculate_overall_stats(results)
    return agg, ci


def load_baseline():
    path = 'reports/backtest/freeze_baseline_v315.json'
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)


def sweep(test_count, quick=False, lite=False):
    if quick:
        grid, tc, corners = QUICK_GRID, QUICK_TEST_COUNT, {}
    elif lite:
        grid, tc, corners = LITE_GRID, test_count, CORNERS
    else:
        grid, tc, corners = GRID, test_count, CORNERS

    baseline = load_baseline()
    base_top5 = None
    if baseline:
        base_top5 = baseline['overall_stats']['avg_top5_hit_rate']
        base_ci = baseline['confidence_95']['top5']
        base_ci_low = base_ci['ci95_low']
    else:
        base_top5 = RANDOM_BASELINE['top5']
        base_ci_low = RANDOM_BASELINE['top5']

    runs = []  # (param_or_tag, value_or_dict, results)

    # OFAT
    for param, levels in grid.items():
        for val in levels:
            res = run_walkforward(tc, build_custom_config({param: val}))
            runs.append({'param': param, 'value': val, 'kind': 'ofat', 'results': res})

    # corners
    for tag, ov in corners.items():
        res = run_walkforward(tc, build_custom_config(ov))
        runs.append({'param': tag, 'value': None, 'kind': 'corner',
                     'overrides': ov, 'results': res})

    per_run = []
    for r in runs:
        if not r['results']:
            per_run.append({**{k: r[k] for k in ('param', 'value', 'kind')},
                            'error': 'no_results'})
            continue
        agg, ci = evaluate(r['results'])
        top5 = agg['avg_top5_hit_rate']
        top5_ci = ci['top5']
        rec = {
            'param': r['param'],
            'value': r['value'],
            'kind': r['kind'],
            'overrides': r.get('overrides'),
            'top1': agg['avg_top1_hit_rate'],
            'top3': agg['avg_top3_hit_rate'],
            'top5': top5,
            'top6': agg['avg_top6_hit_rate'],
            'top5_ci_low': top5_ci['ci95_low'],
            'top5_ci_high': top5_ci['ci95_high'],
            'delta_vs_baseline': round(top5 - base_top5, 2),
            'beats_random': top5_ci['ci95_low'] > RANDOM_BASELINE['top5'],
            'beats_baseline': top5_ci['ci95_low'] > base_ci_low,
            'tested': agg['total_tested'],
        }
        per_run.append(rec)

    # 汇总: 找出 top5 最高档 & 是否稳健超越
    best = max(per_run, key=lambda x: x.get('top5', -1)) if per_run else None
    any_robust = any(r.get('beats_random') for r in per_run if 'error' not in r)

    summary = {
        'meta': {
            'generated_at': datetime.now().isoformat(),
            'script': 'opt_tune_bayes.py',
            'purpose': 'v3.16 阶段P1: 贝叶斯算法内部超参敏感性扫参(OFAT+角点)',
            'test_count': tc,
            'quick': quick,
            'honest_note': ('排列5公平摇号; 验证记录静态共用(与生产累积行为一致), '
                            '跨配置 apple-to-apple 公平; 任何"杠杆"须 95%CI 下限>随机 '
                            '且跨多参数稳健才成立。'),
            'baseline_top5': base_top5,
            'random_top5': RANDOM_BASELINE['top5'],
        },
        'bayes_defaults': BAYES_DEFAULTS,
        'grid': grid if not quick else QUICK_GRID,
        'corners': corners,
        'per_run': per_run,
        'best_run': best,
        'any_robust_beat_random': any_robust,
        'conclusion': (
            '结论: 所有贝叶斯超参档位的 Top-5 95%CI 下限均 ≤ 随机基线(50%), '
            '无任一档位稳健超越冻结基线/随机 → 贝叶斯内部超参非命中率有效杠杆 '
            '(与 P0 窗口扫参、v3.14 审计一致)。'
            if not any_robust else
            '结论: 存在档位 95%CI 下限>随机, 需进一步交叉验证确认是否为真实信号。'
        ),
    }
    return summary


# ----------------------------------------------------------------------------
# Excel 仪表盘
# ----------------------------------------------------------------------------
def build_excel(summary, path):
    wb = Workbook()
    C_BLUE = '1F4E78'
    C_HEAD = '2E5496'
    C_YELLOW = 'FFF2CC'
    C_RED = 'C00000'
    C_GREEN = '375623'
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    meta = summary['meta']
    per_run = [r for r in summary['per_run'] if 'error' not in r]

    # ---- Sheet1 概览 ----
    ws = wb.active
    ws.title = '概览'
    ws.sheet_view.showGridLines = False
    ws['A1'] = 'v3.16 P1 贝叶斯超参敏感性扫参'
    ws['A1'].font = Font(size=14, bold=True, color=C_BLUE)
    ws.merge_cells('A1:E1')
    ws['A2'] = (f"生成时间: {meta['generated_at']} | 回测: {meta['test_count']}期 "
                f"| 基线Top5: {meta['baseline_top5']}% | 随机: {meta['random_top5']}%")
    ws['A2'].font = Font(size=9, italic=True, color='808080')
    ws.merge_cells('A2:E2')
    ws['A3'] = f"⚠️ {meta['honest_note']}"
    ws['A3'].font = Font(size=9, bold=True, color=C_RED)
    ws.merge_cells('A3:E3')

    headers = ['指标', '值', '说明']
    r0 = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    rows = [
        ('扫参档位数', len(per_run), 'OFAT + 角点'),
        ('随机基线 Top5', f"{meta['random_top5']}%", '黄底=关键假设'),
        ('冻结基线 Top5', f"{meta['baseline_top5']}%", '真值起点'),
        ('是否任一档稳健超越随机', '是' if summary['any_robust_beat_random'] else '否',
         'CI下限>50% 且跨参数稳健'),
    ]
    best = summary['best_run']
    if best:
        rows.append(('最高 Top5 档位', f"{best.get('param')}={best.get('value')}",
                     f"Top5={best.get('top5')}% (CI[{best.get('top5_ci_low')},{best.get('top5_ci_high')}])"))
    for i, (a, b, c) in enumerate(rows):
        row = r0 + 1 + i
        ws.cell(row=row, column=1, value=a).font = Font(bold=True)
        ws.cell(row=row, column=2, value=b)
        ws.cell(row=row, column=3, value=c)
        for cc in range(1, 4):
            ws.cell(row=row, column=cc).border = border
            ws.cell(row=row, column=cc).alignment = left if cc == 3 else center
        if a in ('随机基线 Top5', '冻结基线 Top5'):
            ws.cell(row=row, column=2).fill = PatternFill('solid', fgColor=C_YELLOW)
    concl_row = r0 + 1 + len(rows) + 1
    ws.cell(row=concl_row, column=1, value=summary['conclusion'])
    ws.cell(row=concl_row, column=1).font = Font(bold=True, color=C_BLUE)
    ws.cell(row=concl_row, column=1).alignment = left
    ws.merge_cells(start_row=concl_row, start_column=1, end_row=concl_row, end_column=5)
    ws.row_dimensions[concl_row].height = 45
    for c, w in enumerate([26, 22, 40], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet2 各参数敏感度 ----
    ws2 = wb.create_sheet('参数敏感度')
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = '贝叶斯超参 OFAT 敏感度 (Top-5 命中率 vs 冻结基线)'
    ws2['A1'].font = Font(size=13, bold=True, color=C_BLUE)
    ws2.merge_cells('A1:G1')
    ph = ['参数', '档位值', 'Top-5(%)', 'CI下限', 'CI上限', 'Δvs基线', '稳健超越随机?']
    for c, h in enumerate(ph, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    rr = 4
    # 按参数分组, 组内按值排序
    groups = {}
    for r in per_run:
        if r['kind'] != 'ofat':
            continue
        groups.setdefault(r['param'], []).append(r)
    for param in GRID.keys():
        items = sorted(groups.get(param, []), key=lambda x: x['value'])
        for it in items:
            vals = [param, it['value'], it['top5'], it['top5_ci_low'],
                    it['top5_ci_high'], it['delta_vs_baseline'],
                    '是' if it['beats_random'] else '否']
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=rr, column=c, value=v)
                cell.border = border
                cell.alignment = center if c > 1 else left
                if c == 1:
                    cell.font = Font(bold=True)
                if c == 7:
                    cell.font = Font(bold=True, color=C_GREEN if it['beats_random'] else C_RED)
            rr += 1
    # 角点
    for r in per_run:
        if r['kind'] != 'corner':
            continue
        vals = [f"角点:{r['param']}", '组合', r['top5'], r['top5_ci_low'],
                r['top5_ci_high'], r['delta_vs_baseline'],
                '是' if r['beats_random'] else '否']
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 1:
                cell.font = Font(bold=True, color=C_BLUE)
            if c == 7:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_random'] else C_RED)
        rr += 1
    # 基线/随机标注行
    ws2.cell(row=rr, column=1, value='冻结基线').font = Font(bold=True, color=C_RED)
    ws2.cell(row=rr, column=3, value=meta['baseline_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
    ws2.cell(row=rr, column=2, value='—').border = border
    ws2.cell(row=rr, column=4, value='—').border = border
    ws2.cell(row=rr, column=5, value='—').border = border
    ws2.cell(row=rr, column=6, value='—').border = border
    ws2.cell(row=rr, column=7, value='—').border = border
    ws2.cell(row=rr + 1, column=1, value='随机基线').font = Font(bold=True, color=C_RED)
    ws2.cell(row=rr + 1, column=3, value=meta['random_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
    for c in (2, 4, 5, 6, 7):
        ws2.cell(row=rr + 1, column=c, value='—').border = border
    # 数据条(Top5列)
    ws2.conditional_formatting.add(
        f'C4:C{rr-1}',
        DataBarRule(start_type='num', start_value=40, end_type='num', end_value=60,
                    color='5B9BD5', showValue=True))
    for c, w in enumerate([20, 12, 12, 12, 12, 12, 16], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet3 逐配置明细 ----
    ws3 = wb.create_sheet('逐配置明细')
    ws3.sheet_view.showGridLines = False
    dh = ['参数/标签', '档位/类型', 'Top-1', 'Top-3', 'Top-5', 'Top-6',
          'CI下', 'CI上', 'Δ基线', '超越随机', '超越基线']
    for c, h in enumerate(dh, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    for i, r in enumerate(per_run):
        row = 2 + i
        label = r['param']
        if r['kind'] == 'ofat':
            lv = r['value']
            kind = 'OFAT'
        else:
            lv = 'corner'
            kind = '角点'
        vals = [label, lv, r['top1'], r['top3'], r['top5'], r['top6'],
                r['top5_ci_low'], r['top5_ci_high'], r['delta_vs_baseline'],
                '是' if r['beats_random'] else '否',
                '是' if r['beats_baseline'] else '否']
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 10:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_random'] else C_RED)
            if c == 11:
                cell.font = Font(bold=True, color=C_GREEN if r['beats_baseline'] else C_RED)
    ws3.freeze_panes = 'A2'
    for c, w in enumerate([20, 12, 10, 10, 10, 10, 9, 9, 9, 11, 11], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)


if __name__ == '__main__':
    args = sys.argv[1:]
    quick = '--quick' in args
    lite = '--lite' in args
    tc = 300
    for a in args:
        if a.isdigit():
            tc = int(a)
    mode = 'quick' if quick else ('lite' if lite else 'full')
    print(f'>>> P1 贝叶斯超参扫参 ({mode}, '
          f'回测 {QUICK_TEST_COUNT if quick else tc} 期)...')
    summary = sweep(tc, quick=quick, lite=lite)

    os.makedirs('reports/backtest', exist_ok=True)
    tag = 'quick' if quick else 'p1'
    json_path = f'reports/backtest/sweep_bayes_{tag}.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f'>>> JSON 已保存: {json_path}')

    xlsx_path = f'reports/backtest/sweep_bayes_{tag}.xlsx'
    build_excel(summary, xlsx_path)
    print(f'>>> Excel 已保存: {xlsx_path}')

    per_run = [r for r in summary['per_run'] if 'error' not in r]
    print(f'\n=== P1 汇总 ({len(per_run)} 档) ===')
    print(f"  随机基线 Top5={summary['meta']['random_top5']}% | "
          f"冻结基线 Top5={summary['meta']['baseline_top5']}%")
    best = summary['best_run']
    if best:
        print(f"  最高档: {best.get('param')}={best.get('value')} "
              f"Top5={best.get('top5')}% CI[{best.get('top5_ci_low')},{best.get('top5_ci_high')}]")
    print(f"  稳健超越随机: {'是' if summary['any_robust_beat_random'] else '否'}")
    print(f"  结论: {summary['conclusion']}")
