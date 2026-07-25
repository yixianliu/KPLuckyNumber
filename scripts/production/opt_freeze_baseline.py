# -*- coding: utf-8 -*-
# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

"""
v3.16 阶段0: 标准 Top-N 冻结基线生成器

用途: 在 v3.15 封板后, 用七算法融合(P5Predictor, 禁用AI)做 walk-forward 回测,
      生成标准 Top-1/3/5/6 命中率冻结基线 freeze_baseline_v315.json,
      作为后续所有量化调优的唯一真值起点(方案见 reports/diagnostic/排列5_命中率优化方案.md)。

与既有 opt_backtest.py 口径一致(复制 Backtester 评估逻辑), 但:
  - 不 import matplotlib (本托管 venv 缺, 且冻结期不改 backtester.py)
  - 额外产出逐期明细 + 95% 置信区间 + 随机基线对比
  - 输出文件名固定 freeze_baseline_v315.json (不覆盖 opt_baseline.json)

用法:
    python opt_freeze_baseline.py            # 默认回测最近 300 期
    python opt_freeze_baseline.py 500        # 回测最近 500 期
"""
import sys
import json
import logging
from datetime import datetime
from collections import defaultdict

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter

logging.getLogger().setLevel(logging.WARNING)

from modules.predictor import P5Predictor
from modules.database import P5Database

POSITION_NAMES = ['万位', '千位', '百位', '十位', '个位']
RANDOM_BASELINE = {'top1': 10.0, 'top3': 30.0, 'top5': 50.0, 'top6': 60.0}


# ----------------------------------------------------------------------------
# 评估逻辑(复制 Backtester._evaluate_prediction 口径, 增补 top6)
# ----------------------------------------------------------------------------
def _evaluate_prediction(prediction_result, actual_numbers, target_issue):
    fused_probs = prediction_result.get('fused_probabilities', [])
    position_accuracy = []
    for pos in range(5):
        if pos >= len(fused_probs):
            break
        pos_probs = fused_probs[pos]
        actual_num = actual_numbers[pos]
        sorted_nums = sorted(pos_probs.items(), key=lambda x: x[1], reverse=True)
        rank = next((i + 1 for i, (n, _) in enumerate(sorted_nums) if n == actual_num), 10)
        position_accuracy.append({
            'position': pos + 1,
            'position_name': POSITION_NAMES[pos],
            'actual_number': actual_num,
            'predicted_rank': rank,
            'predicted_probability': round(pos_probs.get(actual_num, 0), 6),
            'top1_hit': rank == 1,
            'top3_hit': rank <= 3,
            'top5_hit': rank <= 5,
            'top6_hit': rank <= 6,
        })
    top1_hits = sum(1 for p in position_accuracy if p['top1_hit'])
    top3_hits = sum(1 for p in position_accuracy if p['top3_hit'])
    top5_hits = sum(1 for p in position_accuracy if p['top5_hit'])
    top6_hits = sum(1 for p in position_accuracy if p['top6_hit'])
    overall_score = round((top1_hits * 40 + top3_hits * 20) / 5, 2)
    brier_scores = []
    for pos in range(5):
        if pos >= len(fused_probs):
            break
        prob_hit = fused_probs[pos].get(actual_numbers[pos], 0)
        brier_scores.append((prob_hit - 1) ** 2)
    avg_brier = round(sum(brier_scores) / len(brier_scores), 6) if brier_scores else 1.0
    calibration_score = round(max(0, 1 - avg_brier) * 100, 2)
    return {
        'target_issue': target_issue,
        'actual_numbers': actual_numbers,
        'position_accuracy': position_accuracy,
        'top1_hit_count': top1_hits,
        'top3_hit_count': top3_hits,
        'top5_hit_count': top5_hits,
        'top6_hit_count': top6_hits,
        'overall_score': overall_score,
        'calibration_score': calibration_score,
        'avg_brier_score': avg_brier,
    }


def _calculate_overall_stats(results):
    if not results:
        return {}, {}
    total = len(results)
    avg_overall_score = round(float(np.mean([r['overall_score'] for r in results])), 2)
    avg_top1 = round(float(np.mean([r['top1_hit_count'] for r in results])), 4)
    avg_top3 = round(float(np.mean([r['top3_hit_count'] for r in results])), 4)
    avg_top5 = round(float(np.mean([r['top5_hit_count'] for r in results])), 4)
    avg_top6 = round(float(np.mean([r['top6_hit_count'] for r in results])), 4)
    avg_cal = round(float(np.mean([r['calibration_score'] for r in results])), 2)

    pos_top1 = [0.0] * 5
    pos_top3 = [0.0] * 5
    pos_top5 = [0.0] * 5
    pos_top6 = [0.0] * 5
    for r in results:
        for item in r['position_accuracy']:
            idx = item['position'] - 1
            if 0 <= idx < 5:
                if item['top1_hit']:
                    pos_top1[idx] += 1
                if item['top3_hit']:
                    pos_top3[idx] += 1
                if item['top5_hit']:
                    pos_top5[idx] += 1
                if item['top6_hit']:
                    pos_top6[idx] += 1
    for i in range(5):
        pos_top1[i] = round(pos_top1[i] / total * 100, 2)
        pos_top3[i] = round(pos_top3[i] / total * 100, 2)
        pos_top5[i] = round(pos_top5[i] / total * 100, 2)
        pos_top6[i] = round(pos_top6[i] / total * 100, 2)

    full_match = sum(1 for r in results if r['top1_hit_count'] == 5)
    full_match_rate = round(full_match / total * 100, 2)

    agg = {
        'total_tested': total,
        'avg_overall_score': avg_overall_score,
        'avg_top1_hit_rate': round(avg_top1 / 5 * 100, 2),
        'avg_top3_hit_rate': round(avg_top3 / 5 * 100, 2),
        'avg_top5_hit_rate': round(avg_top5 / 5 * 100, 2),
        'avg_top6_hit_rate': round(avg_top6 / 5 * 100, 2),
        'avg_calibration_score': avg_cal,
        'full_match_count': full_match,
        'full_match_rate': full_match_rate,
        'position_top1_rates': {POSITION_NAMES[i]: pos_top1[i] for i in range(5)},
        'position_top3_rates': {POSITION_NAMES[i]: pos_top3[i] for i in range(5)},
        'position_top5_rates': {POSITION_NAMES[i]: pos_top5[i] for i in range(5)},
        'position_top6_rates': {POSITION_NAMES[i]: pos_top6[i] for i in range(5)},
    }

    # 95% 置信区间(正态近似, 每位命中为伯努利试验, 共 total*5 次)
    ci = {}
    for key, hits_per_period in [('top1', avg_top1), ('top3', avg_top3),
                                 ('top5', avg_top5), ('top6', avg_top6)]:
        p = hits_per_period / 5  # 每位命中比例
        se = (p * (1 - p) / (total * 5)) ** 0.5
        margin = 1.96 * se * 100
        ci[key] = {
            'rate': round(p * 100, 2),
            'ci95_low': round(max(0.0, p * 100 - margin), 2),
            'ci95_high': round(min(100.0, p * 100 + margin), 2),
        }
    return agg, ci


# ----------------------------------------------------------------------------
# 配置指纹(冻结快照, 从 predictor.config 安全提取已知字段)
# ----------------------------------------------------------------------------
def _config_fingerprint(predictor):
    cfg = predictor.config.config
    alg = cfg.get('algorithms', {}) if isinstance(cfg, dict) else {}
    g = cfg.get('global', {}) if isinstance(cfg, dict) else {}

    def w(name):
        return alg.get(name, {}).get('weight') if isinstance(alg.get(name), dict) else None

    return {
        'algo_weights': {
            'frequency_weighted': w('frequency_weighted'),
            'omission_regression': w('omission_regression'),
            'bayesian': w('bayesian'),
            'trend_momentum': w('trend_momentum'),
            'markov': w('markov'),
            'form_continuity': w('form_continuity'),
            'feature_engineering': w('feature_engineering'),
        },
        'global': {
            'enable_adaptive_weights': g.get('enable_adaptive_weights'),
            'enable_boundary_protection': g.get('enable_boundary_protection'),
            'position_top_n': g.get('position_top_n'),
            'adaptive_metric': g.get('adaptive_metric'),
            'ewma_alpha': g.get('ewma_alpha'),
        },
    }


# ----------------------------------------------------------------------------
# 主流程
# ----------------------------------------------------------------------------
def run_baseline(test_count=300):
    p = P5Predictor()
    p.ai_available = False
    p.config.config.setdefault('global', {})['enable_ai_model'] = False

    db = P5Database()
    db.connect()
    history = db.get_history_data(limit=None, order='ASC')
    total_hist = len(history)
    db.disconnect()

    start_index = max(50, total_hist - test_count)
    effective = min(test_count, total_hist - start_index)

    results = []
    for i in range(start_index, start_index + effective):
        train = history[:i]
        target = history[i]['issue']
        actual = history[i]['numbers']
        pr = p.predict(train, target)
        if 'error' in pr:
            continue
        results.append(_evaluate_prediction(pr, actual, target))

    agg, ci = _calculate_overall_stats(results)
    fingerprint = _config_fingerprint(p)

    # 偏差 vs 随机基线
    deviations = {
        k: {
            'rate': agg[f'avg_{k}_hit_rate'],
            'random': RANDOM_BASELINE[k],
            'deviation': round(agg[f'avg_{k}_hit_rate'] - RANDOM_BASELINE[k], 2),
            'ci95_low': ci[k]['ci95_low'],
            'ci95_high': ci[k]['ci95_high'],
            'beats_random': ci[k]['ci95_low'] > RANDOM_BASELINE[k],
        }
        for k in ('top1', 'top3', 'top5', 'top6')
    }

    summary = {
        'meta': {
            'generated_at': datetime.now().isoformat(),
            'script': 'opt_freeze_baseline.py',
            'purpose': 'v3.16 阶段0: 七算法融合标准Top-N冻结基线(后续调优唯一真值起点)',
            'test_count_requested': test_count,
            'test_count_effective': effective,
            'start_index': start_index,
            'total_history': total_hist,
            'ai_disabled': True,
            'honest_note': '排列5公平摇号, 历史走势无法稳定超越随机基线; 本基线为可复现真值, 非"打败随机"承诺',
        },
        'config_fingerprint': fingerprint,
        'overall_stats': agg,
        'confidence_95': ci,
        'random_baseline': RANDOM_BASELINE,
        'deviation_vs_random': deviations,
        'per_period_details': [
            {
                'issue': r['target_issue'],
                'actual': r['actual_numbers'],
                'top1_hits': r['top1_hit_count'],
                'top3_hits': r['top3_hit_count'],
                'top5_hits': r['top5_hit_count'],
                'top6_hits': r['top6_hit_count'],
                'score': r['overall_score'],
                'calibration': r['calibration_score'],
                'ranks': [pa['predicted_rank'] for pa in r['position_accuracy']],
            }
            for r in results
        ],
    }
    return summary


# ----------------------------------------------------------------------------
# Excel 仪表盘
# ----------------------------------------------------------------------------
def build_excel(summary, path):
    wb = Workbook()

    # 配色: 蓝=基线/输入, 黑=公式值, 红=预警, 黄底=关键假设
    C_BLUE = '1F4E78'
    C_HEAD = '2E5496'
    C_YELLOW = 'FFF2CC'
    C_RED = 'C00000'
    C_GREEN = '375623'
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ---- Sheet1 概览 ----
    ws = wb.active
    ws.title = '概览'
    ws.sheet_view.showGridLines = False
    ws['A1'] = '排列5 v3.15 七算法融合 — 标准 Top-N 冻结基线'
    ws['A1'].font = Font(size=14, bold=True, color=C_BLUE)
    ws.merge_cells('A1:F1')

    meta = summary['meta']
    ws['A2'] = f"生成时间: {meta['generated_at']}  |  回测期数: {meta['test_count_effective']}  |  AI: 已禁用"
    ws['A2'].font = Font(size=9, italic=True, color='808080')
    ws.merge_cells('A2:F2')
    ws['A3'] = f"⚠️ {meta['honest_note']}"
    ws['A3'].font = Font(size=9, bold=True, color=C_RED)
    ws.merge_cells('A3:F3')

    # 汇总表
    headers = ['指标', '实测命中率(%)', '随机基线(%)', '偏差', '95%CI 下限', '95%CI 上限', '是否超越随机']
    r0 = 5
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r0, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    dev = summary['deviation_vs_random']
    for i, k in enumerate(('top1', 'top3', 'top5', 'top6')):
        row = r0 + 1 + i
        d = dev[k]
        vals = [f'Top-{k[3:]}', d['rate'], d['random'], d['deviation'],
                d['ci95_low'], d['ci95_high'], '是' if d['beats_random'] else '否(在基线内)']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center if c > 1 else left
            if c == 1:
                cell.font = Font(bold=True)
            if c == 7:
                cell.font = Font(bold=True, color=C_GREEN if d['beats_random'] else C_RED)
            if c == 3:
                cell.fill = PatternFill('solid', fgColor=C_YELLOW)  # 关键假设: 随机基线
    # 完全命中率 + 校准
    agg = summary['overall_stats']
    r_end = r0 + 5
    ws.cell(row=r_end, column=1, value='完全命中率(5/5)').font = Font(bold=True)
    ws.cell(row=r_end, column=2, value=agg['full_match_rate'])
    ws.cell(row=r_end, column=3, value='0.00').fill = PatternFill('solid', fgColor=C_YELLOW)
    ws.cell(row=r_end, column=4, value=agg['full_match_rate'])
    ws.cell(row=r_end + 1, column=1, value='平均综合得分').font = Font(bold=True)
    ws.cell(row=r_end + 1, column=2, value=agg['avg_overall_score'])
    ws.cell(row=r_end + 2, column=1, value='平均概率校准(Brier→100)').font = Font(bold=True)
    ws.cell(row=r_end + 2, column=2, value=agg['avg_calibration_score'])

    # 结论
    concl_row = r_end + 4
    any_beat = any(dev[k]['beats_random'] for k in dev)
    conclusion = ('结论: 所有标准 Top-N 命中率的 95%CI 下限均 ≤ 随机基线, '
                  '即 v3.15 七算法融合未稳定超越随机基线(符合 v3.14 审计结论)。'
                  '本基线作为后续调优的唯一真值起点。') if not any_beat else \
                 ('结论: 存在部分指标 95%CI 下限 > 随机基线, 需多窗口交叉验证确认是否为真实信号。')
    ws.cell(row=concl_row, column=1, value=conclusion)
    ws.cell(row=concl_row, column=1).font = Font(bold=True, color=C_BLUE)
    ws.cell(row=concl_row, column=1).alignment = left
    ws.merge_cells(start_row=concl_row, start_column=1, end_row=concl_row, end_column=7)
    ws.row_dimensions[concl_row].height = 45

    widths = [22, 16, 14, 12, 14, 14, 18]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet2 逐位置命中率 ----
    ws2 = wb.create_sheet('逐位置命中率')
    ws2.sheet_view.showGridLines = False
    ws2['A1'] = '各位置 Top-N 命中率(%)'
    ws2['A1'].font = Font(size=13, bold=True, color=C_BLUE)
    ws2.merge_cells('A1:E1')
    ph = ['位置', 'Top-1', 'Top-3', 'Top-5', 'Top-6']
    for c, h in enumerate(ph, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    rates_keys = [('position_top1_rates', 'top1'), ('position_top3_rates', 'top3'),
                  ('position_top5_rates', 'top5'), ('position_top6_rates', 'top6')]
    for i, name in enumerate(POSITION_NAMES):
        row = 4 + i
        ws2.cell(row=row, column=1, value=name).font = Font(bold=True)
        ws2.cell(row=row, column=1).border = border
        for c, (rk, _) in enumerate(rates_keys, 2):
            v = agg[rk][name]
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center
        # 随机基线标注行
    base_row = 4 + len(POSITION_NAMES)
    ws2.cell(row=base_row, column=1, value='随机基线').font = Font(bold=True, color=C_RED)
    for c, (_, bk) in enumerate(rates_keys, 2):
        cell = ws2.cell(row=base_row, column=c, value=RANDOM_BASELINE[bk])
        cell.fill = PatternFill('solid', fgColor=C_YELLOW)
        cell.border = border
        cell.alignment = center
    # 数据条
    last = get_column_letter(5)
    for c in range(2, 6):
        col = get_column_letter(c)
        ws2.conditional_formatting.add(
            f'{col}4:{col}{base_row-1}',
            DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                        color='5B9BD5', showValue=True))
    for c, w in enumerate([10, 10, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet3 逐期明细 ----
    ws3 = wb.create_sheet('逐期明细')
    ws3.sheet_view.showGridLines = False
    dh = ['期号', '开奖号码', 'Top-1命中数', 'Top-3命中数', 'Top-5命中数', 'Top-6命中数', '得分', '校准', '各位置rank']
    for c, h in enumerate(dh, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=C_HEAD)
        cell.alignment = center
        cell.border = border
    for i, r in enumerate(summary['per_period_details']):
        row = 2 + i
        actual = ''.join(str(x) for x in r['actual'])
        vals = [r['issue'], actual, r['top1_hits'], r['top3_hits'], r['top5_hits'],
                r['top6_hits'], r['score'], r['calibration'], str(r['ranks'])]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.border = border
            cell.alignment = center if c != 2 and c != 9 else left
    ws3.freeze_panes = 'A2'
    for c, w in enumerate([14, 14, 12, 12, 12, 12, 10, 10, 22], 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)


# ----------------------------------------------------------------------------
if __name__ == '__main__':
    tc = int(sys.argv[1]) if len(sys.argv) > 1 else 300
    print(f'>>> 生成 v3.15 冻结基线 (回测最近 {tc} 期, 禁用AI)...')
    summary = run_baseline(tc)

    import os
    os.makedirs('reports/backtest', exist_ok=True)
    json_path = 'reports/backtest/freeze_baseline_v315.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f'>>> JSON 已保存: {json_path}')

    xlsx_path = 'reports/backtest/freeze_baseline_v315.xlsx'
    build_excel(summary, xlsx_path)
    print(f'>>> Excel 已保存: {xlsx_path}')

    agg = summary['overall_stats']
    dev = summary['deviation_vs_random']
    print('\n=== 冻结基线汇总 (标准 Top-N, 最近%d期) ===' % agg['total_tested'])
    for k in ('top1', 'top3', 'top5', 'top6'):
        d = dev[k]
        print(f"  {k}: 实测 {d['rate']:.2f}% | 随机 {d['random']:.1f}% | 偏差 {d['deviation']:+.2f}% "
              f"| 95%CI [{d['ci95_low']}, {d['ci95_high']}] | 超越随机: {'是' if d['beats_random'] else '否'}")
    print(f"  完全命中率: {agg['full_match_rate']}% | 平均得分: {agg['avg_overall_score']} | 校准: {agg['avg_calibration_score']}")
