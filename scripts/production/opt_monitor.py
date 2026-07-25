# -*- coding: utf-8 -*-
"""
v3.16 监控层 (Monitoring Layer) — opt_monitor.py

把冻结基线(opt_freeze_baseline) + 调优框架(tuning_config) + 扫参(opt_tune_sweep)
变成"持续可监控资产"。四大组件:

  1. 滚动漂移检测 (Rolling Drift Detection)
     在完整 walk-forward 的逐期 rank 序列上做分桶 + 滚动窗口统计, 对比
     冻结基线 Top-5 命中率(49.60%, CI[47.07,52.13]) 与随机基线(50%),
     用二项近似 z 检验判定尾部是否显著负向漂移 -> HEALTHY / WATCH / ALERT。

  2. 信号源独立命中率 (Per-Signal Hit Rate)
     对 7 个融合算法分别做 isolation(仅启用该算法, weight=1) 与
     ablation(冻结基线中禁用该算法, 其余重新归一化) walk-forward,
     计算独立命中率与边际贡献, 识别"哪个信号源在贡献 / 在拖后腿"。

  3. 质量 Gate (Quality Gates)
     数据质量 gate: 历史数据缺失率 / 重复率 / 格式合法性。
     性能 gate: 当前配置 walk-forward 的 Top-5 / Top-1 CI 下限 vs 冻结基线 CI 下限。
     任一 FAIL -> 触发告警。

  4. 退化告警 + 回滚 (Degradation Alert & Rollback)
     综合 drift + gate 给出 status 状态机; DEGRADED 时 recommended_action=
     'rollback_to_freeze', 并提供 `python opt_monitor.py rollback` 将
     tuning_config.yaml 重置为控制组模板(=冻结态)。

诚实前提: 排列5 公平摇号, 历史走势无法稳定超越随机; 本层只做"可监控 / 可告警 /
可回滚", 不承诺提升。所有判定基于 95%CI + 二项检验, 避免把噪声当信号。

零改封板核心(gui/pipeline/predictor/database)。纯新增。

用法:
  python opt_monitor.py all [--window 300] [--signals-window 100]
  python opt_monitor.py drift  [--window 300]
  python opt_monitor.py signals [--window 100]
  python opt_monitor.py gate   [--window 300]
  python opt_monitor.py status
  python opt_monitor.py rollback [--force]
"""
from __future__ import annotations

# --- 路径锚定(B3修复): 向上搜索项目根(modules/+main.py), 注入 sys.path ---
import os
import sys
def _find_project_root(_start):
    _cur = os.path.abspath(_start)
    while True:
        if os.path.isdir(os.path.join(_cur, 'modules')) and \
           os.path.isfile(os.path.join(_cur, 'main.py')):
            return _cur
        _p = os.path.dirname(_cur)
        if _p == _cur:
            return os.path.dirname(os.path.abspath(_start))
        _cur = _p
_PROJECT_ROOT = _find_project_root(__file__)
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

import sys
import os
import json
import math
import copy
import logging
from datetime import datetime
from collections import OrderedDict

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logging.getLogger().setLevel(logging.WARNING)

# 复用阶段0评估内核(无 matplotlib 依赖, 口径与 Backtester 一致)
from opt_freeze_baseline import (
    _evaluate_prediction, _calculate_overall_stats, _config_fingerprint,
    POSITION_NAMES, RANDOM_BASELINE,
)
from modules.predictor import P5Predictor, P5PredictorConfig
from modules.database import P5Database
from tuning_config import _control_template, TuningConfig, EXPECTED_ALGOS

FREEZE_BASELINE_PATH = 'reports/backtest/freeze_baseline_v315.json'
MONITOR_DIR = 'reports/monitor'
LATEST_STATUS_PATH = os.path.join(MONITOR_DIR, 'latest_status.json')
TUNING_CONFIG_PATH = 'tuning_config.yaml'

Z95 = 1.96
ALGO_LABELS = OrderedDict([
    ('frequency_weighted', '频率加权'),
    ('omission_regression', '遗漏回归'),
    ('bayesian_inference', '贝叶斯推断'),
    ('trend_momentum', '趋势动量'),
    ('markov_transition', '马尔可夫转移'),
    ('pattern_continuation', '形态延续'),
    ('feature_engineering', '特征工程'),
])


# ----------------------------------------------------------------------------
# 统计工具(零额外依赖, 正态近似)
# ----------------------------------------------------------------------------
def norm_cdf(z: float) -> float:
    """标准正态 CDF(用误差函数近似)"""
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def proportion_ci(hits: int, n: int, z: float = Z95):
    """二项分布比例的正态近似置信区间"""
    if n <= 0:
        return (0.0, 0.0, 0.0)
    p = hits / n
    se = math.sqrt(p * (1.0 - p) / n)
    margin = z * se
    return (p * 100.0, max(0.0, (p - margin) * 100.0), min(100.0, (p + margin) * 100.0))


def binomial_z_test(p_hat: float, n: int, p0: float):
    """单样本二项 z 检验(two-sided)。返回 (z, p_value)。
    p_hat/p0 为比例(0~1)。"""
    if n <= 0 or p0 <= 0 or p0 >= 1:
        return (0.0, 1.0)
    se0 = math.sqrt(p0 * (1.0 - p0) / n)
    if se0 <= 0:
        return (0.0, 1.0)
    z = (p_hat - p0) / se0
    p_value = 2.0 * (1.0 - norm_cdf(abs(z)))
    return (z, p_value)


# ----------------------------------------------------------------------------
# walk-forward 内核(返回逐期明细, 供漂移滚动统计)
# ----------------------------------------------------------------------------
def run_walk_forward_details(predictor, test_count=300, start_index_min=50):
    db = P5Database()
    db.connect()
    history = db.get_history_data(limit=None, order='ASC')
    total_hist = len(history)
    db.disconnect()

    start_index = max(start_index_min, total_hist - test_count)
    effective = min(test_count, total_hist - start_index)

    results = []
    for i in range(start_index, start_index + effective):
        train = history[:i]
        target = history[i]['issue']
        actual = history[i]['numbers']
        pr = predictor.predict(train, target)
        if 'error' in pr:
            continue
        results.append(_evaluate_prediction(pr, actual, target))

    agg, ci = _calculate_overall_stats(results)
    return results, agg, ci, {
        'meta': {'test_count_effective': effective, 'start_index': start_index,
                 'total_history': total_hist},
        'fingerprint': _config_fingerprint(predictor),
    }


def _build_predictor_from_template(modifier):
    """基于控制组模板构造预测器; modifier(tpl) 原地修改配置。禁用 AI。"""
    tpl = _control_template()
    modifier(tpl)
    tc = TuningConfig(tpl)
    ok, errs = tc.validate()
    if not ok:
        raise ValueError(f'配置校验失败: {errs}')
    custom = tc.to_predictor_custom_config()
    cfg = P5PredictorConfig(custom)
    p = P5Predictor(cfg)
    p.ai_available = False
    p.config.config.setdefault('global', {})['enable_ai_model'] = False
    return p


def _isolation_predictor(algo_name):
    def mod(tpl):
        for name, blk in tpl['algorithms'].items():
            if name == algo_name:
                blk['enabled'] = True
                blk['weight'] = 1.0
            else:
                blk['enabled'] = False
                blk['weight'] = 0.0
    return _build_predictor_from_template(mod)


def _ablation_predictor(algo_name):
    def mod(tpl):
        tpl['algorithms'][algo_name]['enabled'] = False
        # 不改 weight: 归一化时自动排除该算法, 体现"移除后的边际贡献"
    return _build_predictor_from_template(mod)


def _current_config_predictor():
    """用当前 tuning_config.yaml(若存在)构造; 否则回退控制组。"""
    if os.path.exists(TUNING_CONFIG_PATH):
        try:
            tc = TuningConfig.load(TUNING_CONFIG_PATH)
            ok, _ = tc.validate()
            if ok:
                custom = tc.to_predictor_custom_config()
                cfg = P5PredictorConfig(custom)
                p = P5Predictor(cfg)
                p.ai_available = False
                p.config.config.setdefault('global', {})['enable_ai_model'] = False
                return p, 'tuning_config.yaml'
        except Exception:
            pass
    return _build_predictor_from_template(lambda t: None), 'control_template'


# ----------------------------------------------------------------------------
# 组件1: 滚动漂移检测
# ----------------------------------------------------------------------------
def drift_detection(results, baseline, window=300, bucket=60, roll=30, step=15):
    """在逐期 rank 序列上做分桶 + 滚动窗口统计, 判定尾部漂移。"""
    base_agg = baseline['overall_stats']
    base_ci = baseline['confidence_95']
    p0_top5 = base_agg['avg_top5_hit_rate'] / 100.0          # 冻结基线 Top-5 比例
    base_top5_ci_low = base_ci['top5']['ci95_low'] / 100.0   # 基线 CI 下限(比例)
    base_top1 = base_agg['avg_top1_hit_rate'] / 100.0

    # 逐期: 该期5个位置中实际号码 rank<=5 的命中数(位置级伯努利)
    per_period_top5_hits = [r['top5_hit_count'] for r in results]
    per_period_top1_hits = [r['top1_hit_count'] for r in results]
    n_periods = len(results)

    # ---- 非重叠分桶 ----
    n_buckets = max(1, n_periods // bucket)
    buckets = []
    for b in range(n_buckets):
        seg = per_period_top5_hits[b * bucket: (b + 1) * bucket]
        seg_n = len(seg) * 5
        seg_hits = sum(seg)
        rate, low, high = proportion_ci(seg_hits, seg_n)
        # 桶 vs 基线 二项 z 检验
        z, pval = binomial_z_test(seg_hits / seg_n if seg_n else 0, seg_n, p0_top5)
        buckets.append({
            'bucket_index': b + 1,
            'periods': len(seg),
            'top5_hit_rate': round(rate, 2),
            'ci95_low': round(low, 2),
            'ci95_high': round(high, 2),
            'baseline_top5': round(p0_top5 * 100, 2),
            'z_vs_baseline': round(z, 3),
            'p_value': round(pval, 4),
            'below_baseline_ci': low < base_top5_ci_low * 100.0,
            'significant_neg': (pval < 0.05 and (seg_hits / seg_n if seg_n else 0) < p0_top5),
        })

    # ---- 滚动窗口(重叠, 用于趋势折线) ----
    rolling = []
    for start in range(0, max(0, n_periods - roll) + 1, step):
        seg = per_period_top5_hits[start: start + roll]
        if len(seg) < roll:
            continue
        seg_n = len(seg) * 5
        seg_hits = sum(seg)
        rate, low, high = proportion_ci(seg_hits, seg_n)
        rolling.append({
            'start_period': start + 1,
            'end_period': start + len(seg),
            'top5_hit_rate': round(rate, 2),
            'ci95_low': round(low, 2),
            'ci95_high': round(high, 2),
        })

    # ---- 漂移判定(基于最近一个完整桶 + 最近 roll 期) ----
    last_bucket = buckets[-1] if buckets else None
    # 最近 roll 期直接统计
    tail = per_period_top5_hits[-roll:] if n_periods >= roll else per_period_top5_hits
    tail_n = len(tail) * 5
    tail_hits = sum(tail)
    tail_rate, tail_low, tail_high = proportion_ci(tail_hits, tail_n)
    tail_z, tail_p = binomial_z_test(tail_hits / tail_n if tail_n else 0, tail_n, p0_top5)

    if tail_p < 0.05 and (tail_hits / tail_n if tail_n else 0) < p0_top5:
        drift_level = 'ALERT'
        drift_reason = (f'最近 {len(tail)} 期 Top-5 命中率 {tail_rate:.2f}% 显著低于冻结基线 '
                        f'{p0_top5*100:.2f}% (z={tail_z:.2f}, p={tail_p:.3f}) -> 尾部退化漂移')
    elif (last_bucket and last_bucket['below_baseline_ci']) or tail_low < base_top5_ci_low * 100.0:
        drift_level = 'WATCH'
        drift_reason = (f'最近 {len(tail)} 期 Top-5 命中率 CI 下限 {tail_low:.2f}% 低于基线 CI 下限 '
                        f'{base_top5_ci_low*100:.2f}%, 但未达显著水平(p={tail_p:.3f}) -> 观察')
    else:
        drift_level = 'HEALTHY'
        drift_reason = (f'最近 {len(tail)} 期 Top-5 命中率 {tail_rate:.2f}% 与冻结基线 '
                        f'{p0_top5*100:.2f}% 一致, 无显著漂移')

    return {
        'window': window,
        'bucket_size': bucket,
        'roll_window': roll,
        'baseline_top5_rate': round(p0_top5 * 100, 2),
        'baseline_top5_ci_low': round(base_top5_ci_low * 100, 2),
        'random_top5': RANDOM_BASELINE['top5'],
        'tail_top5_rate': round(tail_rate, 2),
        'tail_ci95_low': round(tail_low, 2),
        'tail_ci95_high': round(tail_high, 2),
        'tail_z': round(tail_z, 3),
        'tail_p_value': round(tail_p, 4),
        'drift_level': drift_level,
        'drift_reason': drift_reason,
        'buckets': buckets,
        'rolling': rolling,
    }


# ----------------------------------------------------------------------------
# 组件2: 信号源独立命中率
# ----------------------------------------------------------------------------
def signal_analysis(signals_window=100):
    """对 7 算法做 isolation + ablation walk-forward。"""
    print(f'>>> 信号源独立命中率 (isolation + ablation, 窗口={signals_window})')
    rows = []
    for name in EXPECTED_ALGOS:
        print(f'  · {ALGO_LABELS.get(name, name)} ...', end=' ', flush=True)

        # isolation: 仅启用该算法
        iso_p = _isolation_predictor(name)
        _, iso_agg, iso_ci, _ = run_walk_forward_details(iso_p, test_count=signals_window)
        iso_top5 = iso_agg['avg_top5_hit_rate']
        iso_top5_ci_low = iso_ci['top5']['ci95_low']

        # ablation: 冻结基线中禁用该算法
        ab_p = _ablation_predictor(name)
        _, ab_agg, ab_ci, _ = run_walk_forward_details(ab_p, test_count=signals_window)
        ab_top5 = ab_agg['avg_top5_hit_rate']
        ab_top5_ci_low = ab_ci['top5']['ci95_low']

        # 边际贡献 = 基线 - ablation (正值=该算法有正贡献)
        # 用冻结基线 Top-5 (49.60%) 作为基准
        baseline_top5 = 49.60
        marginal = round(baseline_top5 - ab_top5, 2)

        rows.append({
            'algo': name,
            'label': ALGO_LABELS.get(name, name),
            'isolation_top5_rate': round(iso_top5, 2),
            'isolation_top5_ci_low': round(iso_top5_ci_low, 2),
            'ablation_top5_rate': round(ab_top5, 2),
            'ablation_top5_ci_low': round(ab_top5_ci_low, 2),
            'marginal_contribution': marginal,
            'beats_random_iso': iso_ci['top5']['ci95_low'] > RANDOM_BASELINE['top5'],
        })
        print(f"iso={iso_top5:.2f}% abl={ab_top5:.2f}% 边际={marginal:+.2f}%")

    # 按边际贡献排序(正贡献在前)
    ranking = sorted(rows, key=lambda r: r['marginal_contribution'], reverse=True)
    return {
        'signals_window': signals_window,
        'baseline_top5': 49.60,
        'random_top5': RANDOM_BASELINE['top5'],
        'rows': rows,
        'ranking_by_marginal': [r['algo'] for r in ranking],
    }


# ----------------------------------------------------------------------------
# 组件3: 质量 Gate
# ----------------------------------------------------------------------------
def quality_gates(window=300):
    """数据质量 gate + 性能 gate。"""
    db = P5Database()
    db.connect()
    history = db.get_history_data(limit=None, order='ASC')
    db.disconnect()

    total = len(history)
    recent = history[-window:] if window <= total else history
    n = len(recent)

    # 数据完整性检查
    seen_issues = set()
    dup = 0
    missing = 0
    bad_format = 0
    for row in recent:
        iss = row.get('issue')
        if iss in seen_issues:
            dup += 1
        seen_issues.add(iss)
        nums = row.get('numbers')
        if nums is None:
            missing += 1
            continue
        if not (isinstance(nums, (list, tuple)) and len(nums) == 5
                and all(isinstance(x, int) and 0 <= x <= 9 for x in nums)):
            bad_format += 1

    dup_rate = round(dup / n * 100, 2) if n else 0.0
    missing_rate = round(missing / n * 100, 2) if n else 0.0
    bad_fmt_rate = round(bad_format / n * 100, 2) if n else 0.0

    data_gate_pass = (dup_rate <= 2.0 and missing_rate <= 2.0 and bad_fmt_rate == 0.0)

    # 性能 gate: 当前配置 walk-forward 的 Top-5 / Top-1 CI 下限 vs 基线
    predictor, cfg_src = _current_config_predictor()
    results, agg, ci, _ = run_walk_forward_details(predictor, test_count=window)

    # 载入基线
    if os.path.exists(FREEZE_BASELINE_PATH):
        with open(FREEZE_BASELINE_PATH, encoding='utf-8') as f:
            baseline = json.load(f)
        base_ci = baseline['confidence_95']
        base_top5_low = base_ci['top5']['ci95_low']
        base_top1_low = base_ci['top1']['ci95_low']
    else:
        base_top5_low, base_top1_low = RANDOM_BASELINE['top5'], RANDOM_BASELINE['top1']

    cur_top5_low = ci['top5']['ci95_low']
    cur_top1_low = ci['top1']['ci95_low']
    perf_top5_pass = cur_top5_low >= base_top5_low
    perf_top1_pass = cur_top1_low >= base_top1_low

    perf_gate_pass = perf_top5_pass and perf_top1_pass

    return {
        'window': window,
        'config_source': cfg_src,
        'data_quality': {
            'rows_checked': n,
            'duplicate_rate': dup_rate,
            'missing_rate': missing_rate,
            'bad_format_rate': bad_fmt_rate,
            'pass': data_gate_pass,
        },
        'performance_gate': {
            'top5_rate': agg['avg_top5_hit_rate'],
            'top5_ci_low': round(cur_top5_low, 2),
            'baseline_top5_ci_low': round(base_top5_low, 2),
            'top1_rate': agg['avg_top1_hit_rate'],
            'top1_ci_low': round(cur_top1_low, 2),
            'baseline_top1_ci_low': round(base_top1_low, 2),
            'top5_pass': perf_top5_pass,
            'top1_pass': perf_top1_pass,
            'pass': perf_gate_pass,
        },
        'overall_pass': data_gate_pass and perf_gate_pass,
    }


# ----------------------------------------------------------------------------
# 报告汇总 + 状态机
# ----------------------------------------------------------------------------
def build_report(drift=None, signals=None, gates=None, window=300, signals_window=100):
    drift_level = drift['drift_level'] if drift else 'UNKNOWN'
    gate_level = 'PASS' if (gates and gates['overall_pass']) else ('FAIL' if gates else 'UNKNOWN')

    if drift_level == 'ALERT' or gate_level == 'FAIL':
        status = 'DEGRADED'
    elif drift_level == 'WATCH':
        status = 'WATCH'
    else:
        status = 'HEALTHY'

    if status == 'DEGRADED':
        recommended_action = 'rollback_to_freeze'
        action_detail = ('检测到退化/性能 gate 失败, 建议将 tuning_config.yaml 重置为控制组(=冻结态), '
                         '命令: python opt_monitor.py rollback --force')
    elif status == 'WATCH':
        recommended_action = 'observe'
        action_detail = '尾部指标偏弱但未显著, 建议持续观察, 下一期复跑 drift 确认。'
    else:
        recommended_action = 'none'
        action_detail = '各项指标在基线范围内, 无需干预。'

    signals_ranking = signals['ranking_by_marginal'] if signals else []
    top_contributor = signals_ranking[0] if signals_ranking else None
    worst_contributor = signals_ranking[-1] if signals_ranking else None

    report = {
        'meta': {
            'generated_at': datetime.now().isoformat(),
            'script': 'opt_monitor.py',
            'purpose': 'v3.16 监控层: 滚动漂移 + 信号源独立命中率 + 质量gate + 退化告警/回滚',
            'window': window,
            'signals_window': signals_window,
            'honest_note': '排列5公平摇号, 本层只监控/告警/回滚, 不承诺提升; 判定基于95%CI+二项检验',
        },
        'status': status,
        'drift_level': drift_level,
        'gate_level': gate_level,
        'recommended_action': recommended_action,
        'action_detail': action_detail,
        'drift': drift,
        'signals': signals,
        'gates': gates,
    }

    # 最新状态机(供回滚/告警读取)
    latest = {
        'generated_at': report['meta']['generated_at'],
        'status': status,
        'drift_level': drift_level,
        'gate_level': gate_level,
        'recommended_action': recommended_action,
        'action_detail': action_detail,
        'top5_rate': drift['tail_top5_rate'] if drift else None,
        'baseline_top5_ci_low': drift['baseline_top5_ci_low'] if drift else None,
        'signals_top_contributor': top_contributor,
        'signals_worst_contributor': worst_contributor,
        'signals_ranking_by_marginal': signals_ranking,
        'baseline_ref': FREEZE_BASELINE_PATH,
    }
    return report, latest


# ----------------------------------------------------------------------------
# Excel 仪表盘
# ----------------------------------------------------------------------------
def build_excel(report, path):
    wb = Workbook()
    C_BLUE = '1F4E78'
    C_HEAD = '2E5496'
    C_YELLOW = 'FFF2CC'
    C_RED = 'C00000'
    C_GREEN = '375623'
    C_AMBER = 'BF8F00'
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    status = report['status']
    status_color = { 'HEALTHY': C_GREEN, 'WATCH': C_AMBER, 'DEGRADED': C_RED }.get(status, '808080')

    # ---- Sheet1 监控总览 ----
    ws = wb.active
    ws.title = '监控总览'
    ws.sheet_view.showGridLines = False
    ws['A1'] = '排列5 v3.16 监控层 — 系统健康状态'
    ws['A1'].font = Font(size=14, bold=True, color=C_BLUE)
    ws.merge_cells('A1:F1')

    ws['A2'] = f"生成时间: {report['meta']['generated_at']}  |  窗口: {report['meta']['window']}期  |  信号窗: {report['meta']['signals_window']}期"
    ws['A2'].font = Font(size=9, italic=True, color='808080')
    ws.merge_cells('A2:F2')

    # 状态机大卡
    ws['A4'] = '系统状态'
    ws['A4'].font = Font(bold=True, size=11)
    ws['B4'] = status
    ws['B4'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['B4'].fill = PatternFill('solid', fgColor=status_color)
    ws['B4'].alignment = center
    ws.merge_cells('B4:C4')

    overview = [
        ('漂移等级', report['drift_level']),
        ('质量 Gate', report['gate_level']),
        ('建议动作', report['recommended_action']),
        ('动作说明', report['action_detail']),
    ]
    r = 6
    for k, v in overview:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        c.alignment = left
        if k == '漂移等级':
            c.font = Font(bold=True, color={'ALERT': C_RED, 'WATCH': C_AMBER, 'HEALTHY': C_GREEN}.get(v, '000000'))
        if k == '质量 Gate':
            c.font = Font(bold=True, color=C_GREEN if v == 'PASS' else C_RED)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    ws.cell(row=r + 1, column=1, value=f"⚠️ {report['meta']['honest_note']}").font = Font(size=9, bold=True, color=C_RED)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    ws.row_dimensions[r + 1].height = 30

    for c, w in enumerate([14, 18, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet2 滚动漂移 ----
    if report.get('drift'):
        d = report['drift']
        ws2 = wb.create_sheet('滚动漂移')
        ws2.sheet_view.showGridLines = False
        ws2['A1'] = '滚动漂移检测 — 分桶 Top-5 命中率 vs 冻结基线'
        ws2['A1'].font = Font(size=13, bold=True, color=C_BLUE)
        ws2.merge_cells('A1:H1')
        ws2['A2'] = (f"基线 Top-5={d['baseline_top5_rate']}% (CI下限 {d['baseline_top5_ci_low']}%)  |  "
                     f"随机={d['random_top5']}%  |  尾部 {d['tail_top5_rate']}% (CI[{d['tail_ci95_low']},{d['tail_ci95_high']}])")
        ws2['A2'].font = Font(size=9, italic=True, color='808080')
        ws2.merge_cells('A2:H2')

        h = ['桶#', '期数', 'Top-5(%)', 'CI下限', 'CI上限', '基线', 'z值', 'p值', '跌破基线CI?', '显著负向?']
        r0 = 4
        for c, hh in enumerate(h, 1):
            cell = ws2.cell(row=r0, column=c, value=hh)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=C_HEAD)
            cell.alignment = center
            cell.border = border
        for i, b in enumerate(d['buckets']):
            row = r0 + 1 + i
            vals = [b['bucket_index'], b['periods'], b['top5_hit_rate'], b['ci95_low'], b['ci95_high'],
                    b['baseline_top5'], b['z_vs_baseline'], b['p_value'],
                    '是' if b['below_baseline_ci'] else '否', '是' if b['significant_neg'] else '否']
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=c, value=v)
                cell.border = border
                cell.alignment = center if c > 1 else center
                if c == 9:
                    cell.font = Font(bold=True, color=C_RED if b['below_baseline_ci'] else C_GREEN)
                if c == 10:
                    cell.font = Font(bold=True, color=C_RED if b['significant_neg'] else C_GREEN)
        # 随机基线标注
        base_row = r0 + 1 + len(d['buckets'])
        ws2.cell(row=base_row, column=1, value='随机基线').font = Font(bold=True, color=C_RED)
        ws2.cell(row=base_row, column=3, value=d['random_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
        ws2.cell(row=base_row, column=6, value=d['baseline_top5_rate']).fill = PatternFill('solid', fgColor=C_YELLOW)

        # 数据条
        last = r0 + len(d['buckets'])
        ws2.conditional_formatting.add(
            f'C{r0+1}:C{last}',
            DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100, color='5B9BD5', showValue=True))

        # 漂移结论
        concl = base_row + 2
        ws2.cell(row=concl, column=1, value=f"[{d['drift_level']}] {d['drift_reason']}").font = Font(bold=True, color=status_color)
        ws2.merge_cells(start_row=concl, start_column=1, end_row=concl, end_column=10)
        ws2.row_dimensions[concl].height = 45

        for c, w in enumerate([8, 8, 11, 11, 11, 9, 9, 9, 14, 13], 1):
            ws2.column_dimensions[get_column_letter(c)].width = w

        # 滚动窗口折线(Sheet2 第二块)
        rroll = concl + 2
        ws2.cell(row=rroll, column=1, value='滚动窗口 Top-5 命中率(窗口=%d, 步长=%d)' % (d['roll_window'], 15)).font = Font(bold=True, color=C_BLUE)
        rh = ['起始期', '结束期', 'Top-5(%)', 'CI下限', 'CI上限']
        for c, hh in enumerate(rh, 1):
            cell = ws2.cell(row=rroll + 1, column=c, value=hh)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=C_HEAD)
            cell.alignment = center
            cell.border = border
        for i, rr in enumerate(d['rolling']):
            row = rroll + 2 + i
            for c, key in enumerate(['start_period', 'end_period', 'top5_hit_rate', 'ci95_low', 'ci95_high'], 1):
                cell = ws2.cell(row=row, column=c, value=rr[key])
                cell.border = border
                cell.alignment = center
        roll_last = rroll + 1 + len(d['rolling'])
        chart = LineChart()
        chart.title = '滚动 Top-5 命中率趋势'
        chart.y_axis.title = '命中率(%)'
        chart.x_axis.title = '窗口区间'
        data = Reference(ws2, min_col=3, min_row=rroll + 1, max_row=roll_last)
        cats = Reference(ws2, min_col=1, min_row=rroll + 2, max_row=roll_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws2.add_chart(chart, 'G' + str(rroll))

    # ---- Sheet3 信号源独立命中率 ----
    if report.get('signals'):
        s = report['signals']
        ws3 = wb.create_sheet('信号源独立命中率')
        ws3.sheet_view.showGridLines = False
        ws3['A1'] = '信号源独立命中率 — isolation(纯信号) + ablation(边际贡献)'
        ws3['A1'].font = Font(size=13, bold=True, color=C_BLUE)
        ws3.merge_cells('A1:H1')
        ws3['A2'] = (f"信号窗={s['signals_window']}期  |  基线 Top-5={s['baseline_top5']}%  |  "
                     f"随机={s['random_top5']}%  |  排序: 按消融边际贡献(正=有贡献)")
        ws3['A2'].font = Font(size=9, italic=True, color='808080')
        ws3.merge_cells('A2:H2')

        h3 = ['算法', '隔离Top-5(%)', '隔离CI下限', '消融Top-5(%)', '消融CI下限', '边际贡献', '隔离超随机?', '排名']
        r0 = 4
        for c, hh in enumerate(h3, 1):
            cell = ws3.cell(row=r0, column=c, value=hh)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=C_HEAD)
            cell.alignment = center
            cell.border = border
        ranking = s['ranking_by_marginal']
        for i, row_d in enumerate(s['rows']):
            row = r0 + 1 + i
            rank = ranking.index(row_d['algo']) + 1
            vals = [row_d['label'], row_d['isolation_top5_rate'], row_d['isolation_top5_ci_low'],
                    row_d['ablation_top5_rate'], row_d['ablation_top5_ci_low'],
                    row_d['marginal_contribution'], '是' if row_d['beats_random_iso'] else '否', rank]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=c, value=v)
                cell.border = border
                cell.alignment = center if c > 1 else left
                if c == 1:
                    cell.font = Font(bold=True)
                if c == 6:
                    cell.font = Font(bold=True, color=C_GREEN if row_d['marginal_contribution'] > 0 else C_RED)
                if c == 7:
                    cell.font = Font(bold=True, color=C_GREEN if row_d['beats_random_iso'] else C_RED)
        base_row = r0 + 1 + len(s['rows'])
        ws3.cell(row=base_row, column=1, value='基线 Top-5').font = Font(bold=True, color=C_BLUE)
        ws3.cell(row=base_row, column=2, value=s['baseline_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
        ws3.cell(row=base_row, column=4, value=s['baseline_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
        ws3.cell(row=base_row + 1, column=1, value='随机').font = Font(bold=True, color=C_RED)
        ws3.cell(row=base_row + 1, column=2, value=s['random_top5']).fill = PatternFill('solid', fgColor=C_YELLOW)
        for c, w in enumerate([16, 14, 12, 14, 12, 12, 14, 8], 1):
            ws3.column_dimensions[get_column_letter(c)].width = w

    # ---- Sheet4 质量 Gate ----
    if report.get('gates'):
        g = report['gates']
        ws4 = wb.create_sheet('质量Gate')
        ws4.sheet_view.showGridLines = False
        ws4['A1'] = '质量 Gate — 数据完整性 + 性能'
        ws4['A1'].font = Font(size=13, bold=True, color=C_BLUE)
        ws4.merge_cells('A1:E1')

        dq = g['data_quality']
        pg = g['performance_gate']
        rows4 = [
            ('数据质量 Gate', '', ''),
            ('  检查行数', dq['rows_checked'], ''),
            ('  重复率(%)', dq['duplicate_rate'], '≤2% 通过'),
            ('  缺失率(%)', dq['missing_rate'], '≤2% 通过'),
            ('  格式异常率(%)', dq['bad_format_rate'], '=0% 通过'),
            ('  结果', 'PASS' if dq['pass'] else 'FAIL', ''),
            ('性能 Gate', '', ''),
            ('  当前 Top-5(%)', pg['top5_rate'], ''),
            ('  当前 Top-5 CI下限', pg['top5_ci_low'], f"≥基线 {pg['baseline_top5_ci_low']} 通过"),
            ('  当前 Top-1(%)', pg['top1_rate'], ''),
            ('  当前 Top-1 CI下限', pg['top1_ci_low'], f"≥基线 {pg['baseline_top1_ci_low']} 通过"),
            ('  结果', 'PASS' if pg['pass'] else 'FAIL', ''),
            ('综合', 'PASS' if g['overall_pass'] else 'FAIL', ''),
        ]
        r0 = 3
        for i, (k, v, note) in enumerate(rows4):
            row = r0 + i
            kc = ws4.cell(row=row, column=1, value=k)
            vc = ws4.cell(row=row, column=2, value=v)
            nc = ws4.cell(row=row, column=3, value=note)
            if k in ('数据质量 Gate', '性能 Gate', '综合'):
                kc.font = Font(bold=True, color=C_BLUE)
                vc.font = Font(bold=True, color=C_GREEN if v == 'PASS' else (C_RED if v == 'FAIL' else '000000'))
            else:
                kc.font = Font(bold=False)
                if v in ('PASS', 'FAIL'):
                    vc.font = Font(bold=True, color=C_GREEN if v == 'PASS' else C_RED)
            for c in (1, 2, 3):
                ws4.cell(row=row, column=c).border = border
        for c, w in enumerate([22, 16, 24], 1):
            ws4.column_dimensions[get_column_letter(c)].width = w

    wb.save(path)


# ----------------------------------------------------------------------------
# 输出
# ----------------------------------------------------------------------------
def _stamp():
    return datetime.now().strftime('%Y%m%d_%H%M%S')


def write_outputs(report, latest, prefix='monitor'):
    os.makedirs(MONITOR_DIR, exist_ok=True)
    stamp = _stamp()
    json_path = os.path.join(MONITOR_DIR, f'{prefix}_{stamp}.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f'>>> JSON 已保存: {json_path}')

    xlsx_path = os.path.join(MONITOR_DIR, f'{prefix}_{stamp}.xlsx')
    build_excel(report, xlsx_path)
    print(f'>>> Excel 已保存: {xlsx_path}')

    with open(LATEST_STATUS_PATH, 'w', encoding='utf-8') as f:
        json.dump(latest, f, ensure_ascii=False, indent=2)
    print(f'>>> 最新状态已保存: {LATEST_STATUS_PATH}')
    return json_path, xlsx_path


# ----------------------------------------------------------------------------
# 子命令
# ----------------------------------------------------------------------------
def cmd_all(args):
    if not os.path.exists(FREEZE_BASELINE_PATH):
        print(f'[WARN] 缺少阶段0基线 {FREEZE_BASELINE_PATH}, 漂移/信号对比将仅对随机基线')
    baseline = None
    if os.path.exists(FREEZE_BASELINE_PATH):
        with open(FREEZE_BASELINE_PATH, encoding='utf-8') as f:
            baseline = json.load(f)

    # drift (基于当前配置 walk-forward 的逐期明细)
    drift = None
    if baseline:
        results, _, _, _ = _current_walkforward(args.window)
        drift = drift_detection(results, baseline, window=args.window)

    signals = signal_analysis(signals_window=args.signals_window) if not args.skip_signals else None
    gates = quality_gates(window=args.window)
    report, latest = build_report(drift, signals, gates, window=args.window, signals_window=args.signals_window)
    write_outputs(report, latest)
    _print_summary(report)


def _current_walkforward(window):
    predictor, _ = _current_config_predictor()
    return run_walk_forward_details(predictor, test_count=window)


def cmd_drift(args):
    if not os.path.exists(FREEZE_BASELINE_PATH):
        print(f'[FAIL] 缺少阶段0基线 {FREEZE_BASELINE_PATH}, 请先运行 opt_freeze_baseline.py')
        return 1
    with open(FREEZE_BASELINE_PATH, encoding='utf-8') as f:
        baseline = json.load(f)
    print(f'>>> 滚动漂移检测 (窗口={args.window})')
    results, _, _, _ = _current_walkforward(args.window)
    drift = drift_detection(results, baseline, window=args.window)
    report, latest = build_report(drift, None, None, window=args.window)
    write_outputs(report, latest, prefix='drift')
    _print_summary(report)
    return 0


def cmd_signals(args):
    signals = signal_analysis(signals_window=args.window)
    report, latest = build_report(None, signals, None, signals_window=args.window)
    write_outputs(report, latest, prefix='signals')
    _print_summary(report)
    return 0


def cmd_gate(args):
    gates = quality_gates(window=args.window)
    report, latest = build_report(None, None, gates, window=args.window)
    write_outputs(report, latest, prefix='gate')
    _print_summary(report)
    return 0


def cmd_status(args):
    if not os.path.exists(LATEST_STATUS_PATH):
        print(f'[FAIL] 无最新状态文件 {LATEST_STATUS_PATH}, 请先运行 opt_monitor.py all/drift')
        return 1
    with open(LATEST_STATUS_PATH, encoding='utf-8') as f:
        latest = json.load(f)
    print('=== 最新监控状态 ===')
    print(f"  状态:      {latest['status']}")
    print(f"  漂移等级:  {latest['drift_level']}")
    print(f"  质量Gate:  {latest['gate_level']}")
    print(f"  建议动作:  {latest['recommended_action']}")
    print(f"  说明:      {latest['action_detail']}")
    if latest.get('top5_rate') is not None:
        print(f"  尾部Top-5: {latest['top5_rate']}% (基线CI下限 {latest['baseline_top5_ci_low']}%)")
    if latest.get('signals_ranking_by_marginal'):
        print(f"  信号贡献排名(正→负): {latest['signals_ranking_by_marginal']}")
    return 0


def cmd_rollback(args):
    if not args.force:
        print(f'[DRY-RUN] 将把 {TUNING_CONFIG_PATH} 重置为控制组模板(=v3.15冻结态)。')
        print(f'          这将覆盖现有调优实验配置(如有)。加 --force 真正执行。')
        return 0
    tpl = _control_template()
    with open(TUNING_CONFIG_PATH, 'w', encoding='utf-8') as f:
        import yaml
        yaml.safe_dump(tpl, f, allow_unicode=True, sort_keys=False)
    print(f'[OK] 已回滚 {TUNING_CONFIG_PATH} -> 控制组(冻结态)。下一步可重新派生调优实验。')
    return 0


def _print_summary(report):
    print('\n=== 监控层汇总 ===')
    print(f"  系统状态: {report['status']}  (漂移={report['drift_level']}, Gate={report['gate_level']})")
    print(f"  建议动作: {report['recommended_action']}")
    print(f"  说明:     {report['action_detail']}")
    if report.get('drift'):
        d = report['drift']
        print(f"  尾部Top-5: {d['tail_top5_rate']}% (CI[{d['tail_ci95_low']},{d['tail_ci95_high']}]) "
              f"vs 基线 {d['baseline_top5_rate']}% (p={d['tail_p_value']})")
    if report.get('signals'):
        s = report['signals']
        top = s['ranking_by_marginal'][0]
        bot = s['ranking_by_marginal'][-1]
        print(f"  信号源贡献最高: {ALGO_LABELS.get(top, top)}; 最低: {ALGO_LABELS.get(bot, bot)}")


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------
def main():
    import argparse
    p = argparse.ArgumentParser(description='v3.16 监控层 (opt_monitor.py)')
    sub = p.add_subparsers(dest='cmd', required=True)

    a = sub.add_parser('all', help='运行全部组件(漂移+信号+gate)')
    a.add_argument('--window', type=int, default=300, help='漂移/性能gate窗口(期)')
    a.add_argument('--signals-window', type=int, default=100, help='信号源分析窗口(期)')
    a.add_argument('--skip-signals', action='store_true', help='跳过较慢的信号源分析')
    a.set_defaults(func=cmd_all)

    d = sub.add_parser('drift', help='仅滚动漂移检测')
    d.add_argument('--window', type=int, default=300)
    d.set_defaults(func=cmd_drift)

    sg = sub.add_parser('signals', help='仅信号源独立命中率')
    sg.add_argument('--window', type=int, default=100, help='信号源分析窗口(期)')
    sg.set_defaults(func=cmd_signals)

    g = sub.add_parser('gate', help='仅质量 gate')
    g.add_argument('--window', type=int, default=300)
    g.set_defaults(func=cmd_gate)

    st = sub.add_parser('status', help='打印最新监控状态')
    st.set_defaults(func=cmd_status)

    rb = sub.add_parser('rollback', help='回滚 tuning_config.yaml 为控制组(冻结态)')
    rb.add_argument('--force', action='store_true', help='真正写文件(否则 dry-run)')
    rb.set_defaults(func=cmd_rollback)

    args = p.parse_args()
    return args.func(args) or 0


if __name__ == '__main__':
    raise SystemExit(main())
